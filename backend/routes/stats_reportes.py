"""Router for stats, reportes, kardex, backup and export endpoints."""
import json
import uuid
import io
from datetime import datetime, timezone, date
from fastapi import APIRouter, HTTPException, Depends, Query, UploadFile, File, Request
from fastapi.responses import StreamingResponse
from db import get_pool
from auth_utils import get_current_user
from decimal import Decimal
from helpers import row_to_dict, parse_jsonb, registrar_actividad
from typing import Optional, List
from pydantic import BaseModel
from models import ESTADOS_PRODUCCION

router = APIRouter(prefix="/api")

@router.get("/stats")
async def get_stats():
    pool = await get_pool()
    async with pool.acquire() as conn:
        marcas = await conn.fetchval("SELECT COUNT(*) FROM prod_marcas")
        tipos = await conn.fetchval("SELECT COUNT(*) FROM prod_tipos")
        entalles = await conn.fetchval("SELECT COUNT(*) FROM prod_entalles")
        telas = await conn.fetchval("SELECT COUNT(*) FROM prod_telas")
        hilos = await conn.fetchval("SELECT COUNT(*) FROM prod_hilos")
        modelos = await conn.fetchval("SELECT COUNT(*) FROM prod_modelos")
        registros = await conn.fetchval("SELECT COUNT(*) FROM prod_registros")
        registros_urgentes = await conn.fetchval("SELECT COUNT(*) FROM prod_registros WHERE urgente = true")
        tallas = await conn.fetchval("SELECT COUNT(*) FROM prod_tallas_catalogo")
        colores = await conn.fetchval("SELECT COUNT(*) FROM prod_colores_catalogo")
        inventario = await conn.fetchval("SELECT COUNT(*) FROM prod_inventario")
        ingresos = await conn.fetchval("SELECT COUNT(*) FROM prod_inventario_ingresos")
        salidas = await conn.fetchval("SELECT COUNT(*) FROM prod_inventario_salidas")
        ajustes = await conn.fetchval("SELECT COUNT(*) FROM prod_inventario_ajustes")
        
        # Alertas de stock: items con stock_minimo > 0 y stock por debajo
        stock_bajo_count = await conn.fetchval("""
            SELECT COUNT(*) FROM prod_inventario 
            WHERE stock_minimo > 0 
              AND stock_actual > 0 
              AND stock_actual <= stock_minimo 
              AND COALESCE(ignorar_alerta_stock, false) = false
        """)
        sin_stock_count = await conn.fetchval("""
            SELECT COUNT(*) FROM prod_inventario 
            WHERE stock_minimo > 0 
              AND stock_actual <= 0 
              AND COALESCE(ignorar_alerta_stock, false) = false
        """)
        
        estados_count = {}
        for estado in ESTADOS_PRODUCCION:
            count = await conn.fetchval("SELECT COUNT(*) FROM prod_registros WHERE estado = $1", estado)
            estados_count[estado] = count
        
        return {
            "marcas": marcas, "tipos": tipos, "entalles": entalles, "telas": telas, "hilos": hilos,
            "modelos": modelos, "registros": registros, "registros_urgentes": registros_urgentes,
            "tallas": tallas, "colores": colores, "inventario": inventario,
            "ingresos_count": ingresos, "salidas_count": salidas, "ajustes_count": ajustes,
            "estados_count": estados_count,
            "stock_bajo": stock_bajo_count or 0,
            "sin_stock": sin_stock_count or 0,
            "alertas_stock_total": (stock_bajo_count or 0) + (sin_stock_count or 0),
        }

@router.get("/stats/charts")
async def get_stats_charts():
    """Datos para gráficos del dashboard"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        # Registros por marca
        marcas_query = """
            SELECT COALESCE(ma.nombre, 'Sin Marca') as name, COUNT(*) as value
            FROM prod_registros r
            LEFT JOIN prod_modelos m ON r.modelo_id = m.id
            LEFT JOIN prod_marcas ma ON m.marca_id = ma.id
            GROUP BY ma.nombre
            ORDER BY value DESC
            LIMIT 8
        """
        marcas_rows = await conn.fetch(marcas_query)
        registros_por_marca = [{"name": r["name"], "value": r["value"]} for r in marcas_rows]
        
        # Producción mensual (últimos 6 meses)
        mensual_query = """
            SELECT 
                TO_CHAR(fecha_creacion, 'Mon') as mes,
                EXTRACT(MONTH FROM fecha_creacion) as mes_num,
                COUNT(*) as registros
            FROM prod_registros
            WHERE fecha_creacion >= CURRENT_DATE - INTERVAL '6 months'
            GROUP BY TO_CHAR(fecha_creacion, 'Mon'), EXTRACT(MONTH FROM fecha_creacion)
            ORDER BY mes_num
        """
        mensual_rows = await conn.fetch(mensual_query)
        produccion_mensual = [{"mes": r["mes"], "registros": r["registros"]} for r in mensual_rows]
        
        # Registros por tipo
        tipos_query = """
            SELECT COALESCE(t.nombre, 'Sin Tipo') as name, COUNT(*) as value
            FROM prod_registros r
            LEFT JOIN prod_modelos m ON r.modelo_id = m.id
            LEFT JOIN prod_tipos t ON m.tipo_id = t.id
            GROUP BY t.nombre
            ORDER BY value DESC
            LIMIT 8
        """
        tipos_rows = await conn.fetch(tipos_query)
        registros_por_tipo = [{"name": r["name"], "value": r["value"]} for r in tipos_rows]
        
        return {
            "registros_por_marca": registros_por_marca,
            "registros_por_tipo": registros_por_tipo,
            "produccion_mensual": produccion_mensual
        }

# ==================== REPORTE MERMAS ====================

@router.get("/reportes/mermas")
async def get_reporte_mermas(fecha_inicio: str = None, fecha_fin: str = None, persona_id: str = None, servicio_id: str = None):
    """Reporte de mermas por período con totales y estadísticas"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        # Query base con filtros
        query = """
            SELECT m.*, 
                   r.n_corte,
                   p.nombre as persona_nombre,
                   s.nombre as servicio_nombre
            FROM prod_mermas m
            LEFT JOIN prod_registros r ON m.registro_id = r.id
            LEFT JOIN prod_personas_produccion p ON m.persona_id = p.id
            LEFT JOIN prod_servicios_produccion s ON m.servicio_id = s.id
            WHERE 1=1
        """
        params = []
        
        if fecha_inicio:
            params.append(fecha_inicio)
            query += f" AND m.fecha >= ${len(params)}::date"
        if fecha_fin:
            params.append(fecha_fin)
            query += f" AND m.fecha <= ${len(params)}::date"
        if persona_id:
            params.append(persona_id)
            query += f" AND m.persona_id = ${len(params)}"
        if servicio_id:
            params.append(servicio_id)
            query += f" AND m.servicio_id = ${len(params)}"
        
        query += " ORDER BY m.fecha DESC"
        
        rows = await conn.fetch(query, *params)
        mermas = [row_to_dict(r) for r in rows]
        
        # Totales
        total_cantidad = sum(m.get('cantidad', 0) or 0 for m in mermas)
        
        # Mermas por persona
        mermas_por_persona = {}
        for m in mermas:
            persona = m.get('persona_nombre') or 'Sin asignar'
            if persona not in mermas_por_persona:
                mermas_por_persona[persona] = 0
            mermas_por_persona[persona] += m.get('cantidad', 0) or 0
        
        # Mermas por servicio
        mermas_por_servicio = {}
        for m in mermas:
            servicio = m.get('servicio_nombre') or 'Sin servicio'
            if servicio not in mermas_por_servicio:
                mermas_por_servicio[servicio] = 0
            mermas_por_servicio[servicio] += m.get('cantidad', 0) or 0
        
        # Mermas por mes
        mermas_por_mes = {}
        for m in mermas:
            if m.get('fecha'):
                mes = m['fecha'].strftime('%Y-%m') if hasattr(m['fecha'], 'strftime') else str(m['fecha'])[:7]
                if mes not in mermas_por_mes:
                    mermas_por_mes[mes] = 0
                mermas_por_mes[mes] += m.get('cantidad', 0) or 0
        
        return {
            "mermas": mermas,
            "total_registros": len(mermas),
            "total_cantidad": total_cantidad,
            "por_persona": [{"name": k, "value": v} for k, v in mermas_por_persona.items()],
            "por_servicio": [{"name": k, "value": v} for k, v in mermas_por_servicio.items()],
            "por_mes": [{"mes": k, "cantidad": v} for k, v in sorted(mermas_por_mes.items())]
        }

# ==================== REPORTE PRODUCTIVIDAD ====================

@router.get("/reportes/productividad")
async def get_reporte_productividad(fecha_inicio: str = None, fecha_fin: str = None, servicio_id: str = None, persona_id: str = None):
    pool = await get_pool()
    async with pool.acquire() as conn:
        query = "SELECT * FROM prod_movimientos_produccion WHERE fecha_fin IS NOT NULL"
        params = []
        if fecha_inicio:
            params.append(fecha_inicio)
            query += f" AND fecha_fin >= ${len(params)}::date"
        if fecha_fin:
            params.append(fecha_fin)
            query += f" AND fecha_fin <= ${len(params)}::date"
        if servicio_id:
            params.append(servicio_id)
            query += f" AND servicio_id = ${len(params)}"
        if persona_id:
            params.append(persona_id)
            query += f" AND persona_id = ${len(params)}"


        
        rows = await conn.fetch(query, *params)
        
        por_servicio = {}
        por_persona = {}
        
        for m in rows:
            srv_id = m['servicio_id']
            per_id = m['persona_id']
            cantidad = m['cantidad_recibida'] or 0
            costo = float(m['costo_calculado'] or 0)
            
            srv = await conn.fetchrow("SELECT nombre FROM prod_servicios_produccion WHERE id = $1", srv_id)
            srv_nombre = srv['nombre'] if srv else 'Desconocido'
            
            if srv_id not in por_servicio:
                por_servicio[srv_id] = {"servicio_id": srv_id, "servicio_nombre": srv_nombre, "total_cantidad": 0, "total_costo": 0, "movimientos": 0}
            por_servicio[srv_id]['total_cantidad'] += cantidad
            por_servicio[srv_id]['total_costo'] += costo
            por_servicio[srv_id]['movimientos'] += 1
            
            per = await conn.fetchrow("SELECT nombre FROM prod_personas_produccion WHERE id = $1", per_id)
            per_nombre = per['nombre'] if per else 'Desconocido'
            
            if per_id not in por_persona:
                por_persona[per_id] = {"persona_id": per_id, "persona_nombre": per_nombre, "total_cantidad": 0, "total_costo": 0, "movimientos": 0}
            por_persona[per_id]['total_cantidad'] += cantidad
            por_persona[per_id]['total_costo'] += costo
            por_persona[per_id]['movimientos'] += 1
        
        return {
            "por_servicio": list(por_servicio.values()),
            "por_persona": list(por_persona.values()),
            "total_movimientos": len(rows)
        }

# ==================== ENDPOINTS KARDEX E INVENTARIO MOVIMIENTOS ====================

@router.get("/inventario-movimientos")
async def get_inventario_movimientos(
    item_id: str = None,
    tipo: str = None,
    fecha_inicio: str = None, fecha_fin: str = None,
    fecha_desde: str = None, fecha_hasta: str = None,
    limit: int = 100, offset: int = 0,
):
    # Compatibilidad de nombres de params
    f_inicio = fecha_inicio or fecha_desde
    f_fin = fecha_fin or fecha_hasta
    
    pool = await get_pool()
    async with pool.acquire() as conn:
        # Construir queries con JOINs (eliminando N+1)
        unions = []
        params = []
        param_idx = 0
        
        # --- INGRESOS ---
        if not tipo or tipo == 'ingreso':
            ing_where = []
            if item_id:
                param_idx += 1; params.append(item_id)
                ing_where.append(f"i.item_id = ${param_idx}")
            if f_inicio:
                param_idx += 1; params.append(f_inicio)
                ing_where.append(f"i.fecha >= ${param_idx}::timestamp")
            if f_fin:
                param_idx += 1; params.append(f_fin)
                ing_where.append(f"i.fecha <= ${param_idx}::timestamp")
            where_clause = (" AND " + " AND ".join(ing_where)) if ing_where else ""
            unions.append(f"""
                SELECT i.id, 'ingreso' as tipo, i.item_id, inv.nombre as item_nombre, inv.codigo as item_codigo,
                       i.cantidad::float, i.costo_unitario::float, (i.cantidad * i.costo_unitario)::float as costo_total,
                       i.fecha, i.proveedor, i.numero_documento, i.observaciones,
                       NULL as registro_id, NULL as registro_n_corte, NULL as motivo
                FROM prod_inventario_ingresos i
                LEFT JOIN prod_inventario inv ON inv.id = i.item_id
                WHERE 1=1 {where_clause}
            """)
        
        # --- SALIDAS ---
        if not tipo or tipo == 'salida':
            sal_where = []
            if item_id:
                param_idx += 1; params.append(item_id)
                sal_where.append(f"s.item_id = ${param_idx}")
            if f_inicio:
                param_idx += 1; params.append(f_inicio)
                sal_where.append(f"s.fecha >= ${param_idx}::timestamp")
            if f_fin:
                param_idx += 1; params.append(f_fin)
                sal_where.append(f"s.fecha <= ${param_idx}::timestamp")
            where_clause = (" AND " + " AND ".join(sal_where)) if sal_where else ""
            unions.append(f"""
                SELECT s.id, 'salida' as tipo, s.item_id, inv.nombre as item_nombre, inv.codigo as item_codigo,
                       s.cantidad::float, 0::float as costo_unitario, s.costo_total::float,
                       s.fecha, NULL as proveedor, NULL as numero_documento, s.observaciones,
                       s.registro_id, r.n_corte as registro_n_corte, NULL as motivo
                FROM prod_inventario_salidas s
                LEFT JOIN prod_inventario inv ON inv.id = s.item_id
                LEFT JOIN prod_registros r ON r.id = s.registro_id
                WHERE 1=1 {where_clause}
            """)
        
        # --- AJUSTES ---
        if not tipo or tipo.startswith('ajuste'):
            aj_where = []
            if item_id:
                param_idx += 1; params.append(item_id)
                aj_where.append(f"a.item_id = ${param_idx}")
            if f_inicio:
                param_idx += 1; params.append(f_inicio)
                aj_where.append(f"a.fecha >= ${param_idx}::timestamp")
            if f_fin:
                param_idx += 1; params.append(f_fin)
                aj_where.append(f"a.fecha <= ${param_idx}::timestamp")
            where_clause = (" AND " + " AND ".join(aj_where)) if aj_where else ""
            unions.append(f"""
                SELECT a.id, ('ajuste_' || a.tipo) as tipo, a.item_id, inv.nombre as item_nombre, inv.codigo as item_codigo,
                       a.cantidad::float, 0::float as costo_unitario, 0::float as costo_total,
                       a.fecha, NULL as proveedor, NULL as numero_documento, a.observaciones,
                       NULL as registro_id, NULL as registro_n_corte, a.motivo
                FROM prod_inventario_ajustes a
                LEFT JOIN prod_inventario inv ON inv.id = a.item_id
                WHERE 1=1 {where_clause}
            """)
        
        if not unions:
            return []
        
        full_query = " UNION ALL ".join(unions)
        
        # Count total
        count_query = f"SELECT COUNT(*) FROM ({full_query}) sub"
        total = await conn.fetchval(count_query, *params)
        
        # Paginated results
        param_idx += 1; params.append(limit)
        param_idx += 1; params.append(offset)
        data_query = f"SELECT * FROM ({full_query}) sub ORDER BY fecha DESC NULLS LAST LIMIT ${param_idx - 1} OFFSET ${param_idx}"
        rows = await conn.fetch(data_query, *params)
        
        movimientos = [row_to_dict(r) for r in rows]
        return {"items": movimientos, "total": total}

@router.get("/inventario-kardex/{item_id}")
async def get_inventario_kardex_by_path(item_id: str, excluir_migracion: bool = False):
    return await _get_kardex(item_id, excluir_migracion=excluir_migracion)

@router.get("/inventario-kardex")
async def get_inventario_kardex(item_id: str, excluir_migracion: bool = False):
    return await _get_kardex(item_id, excluir_migracion=excluir_migracion)

async def _get_kardex(item_id: str, excluir_migracion: bool = False):
    pool = await get_pool()
    async with pool.acquire() as conn:
        item = await conn.fetchrow("SELECT * FROM prod_inventario WHERE id = $1", item_id)
        if not item:
            raise HTTPException(status_code=404, detail="Item no encontrado")

        movimientos = []

        # Ingresos (no filtrar — los ingresos normales nunca son de migración)
        ingresos_query = "SELECT * FROM prod_inventario_ingresos WHERE item_id = $1"
        ingresos = await conn.fetch(ingresos_query, item_id)
        for ing in ingresos:
            movimientos.append({
                "id": ing['id'],
                "tipo": "ingreso",
                "fecha": ing['fecha'],
                "cantidad": float(ing['cantidad']),
                "costo_unitario": float(ing['costo_unitario']),
                "costo_total": float(ing['cantidad']) * float(ing['costo_unitario']),
                "proveedor": ing['proveedor'],
                "numero_documento": ing['numero_documento'],
                "observaciones": ing['observaciones']
            })

        # Salidas — si excluir_migracion, ocultar las que fueron revertidas por un período de carga inicial
        if excluir_migracion:
            salidas = await conn.fetch(
                """SELECT * FROM prod_inventario_salidas
                   WHERE item_id = $1
                     AND revertida_por_migracion_id IS NULL""",
                item_id
            )
        else:
            salidas = await conn.fetch("SELECT * FROM prod_inventario_salidas WHERE item_id = $1", item_id)
        for sal in salidas:
            registro = None
            modelo_nombre = None
            if sal['registro_id']:
                registro = await conn.fetchrow("""
                    SELECT r.n_corte, COALESCE(m.nombre, r.modelo_manual->>'nombre_modelo') as modelo_nombre
                    FROM prod_registros r
                    LEFT JOIN prod_modelos m ON r.modelo_id = m.id
                    WHERE r.id = $1
                """, sal['registro_id'])
                if registro:
                    modelo_nombre = registro['modelo_nombre']
            sal_cant = float(sal['cantidad'])
            sal_costo = float(sal['costo_total'] or 0)
            movimientos.append({
                "id": sal['id'],
                "tipo": "salida",
                "fecha": sal['fecha'],
                "cantidad": -sal_cant,
                "costo_unitario": round(sal_costo / sal_cant, 4) if sal_cant > 0 else 0,
                "costo_total": sal_costo,
                "registro_id": sal['registro_id'],
                "registro_n_corte": registro['n_corte'] if registro else None,
                "modelo_nombre": modelo_nombre,
                "rollo_id": sal.get('rollo_id'),
            })

        # Ajustes
        ajustes_query = "SELECT * FROM prod_inventario_ajustes WHERE item_id = $1"
        if excluir_migracion:
            ajustes_query += " AND (subtipo IS NULL OR subtipo != 'ajuste_migracion')"
        ajustes = await conn.fetch(ajustes_query, item_id)
        for aj in ajustes:
            cantidad = float(aj['cantidad']) if aj['tipo'] == 'entrada' else -float(aj['cantidad'])
            costo_aj = float(aj['costo_total'] or 0)
            movimientos.append({
                "id": aj['id'],
                "tipo": f"ajuste_{aj['tipo']}",
                "fecha": aj['fecha'],
                "cantidad": cantidad,
                "costo_unitario": round(costo_aj / abs(cantidad), 4) if cantidad else 0,
                "costo_total": costo_aj,
                "motivo": aj['motivo'],
                "observaciones": aj['observaciones']
            })
        
        # Ordenar por fecha
        movimientos.sort(key=lambda x: x['fecha'] if x['fecha'] else datetime.min)
        
        # Calcular saldo acumulado y saldo valorizado
        saldo = 0
        saldo_valor = 0
        for mov in movimientos:
            if mov['tipo'] == 'ingreso':
                saldo += mov['cantidad']
                saldo_valor += mov['costo_total']
            elif mov['tipo'] == 'salida':
                saldo += mov['cantidad']  # ya es negativo
                saldo_valor -= mov['costo_total']
            elif mov['tipo'] == 'ajuste_entrada':
                saldo += abs(mov['cantidad'])
                saldo_valor += mov['costo_total']
            elif mov['tipo'] == 'ajuste_salida':
                saldo -= abs(mov['cantidad'])
                saldo_valor -= mov['costo_total']
            mov['saldo'] = saldo
            mov['saldo_valorizado'] = round(saldo_valor, 2)

        # Valorizado actual (capas FIFO vivas)
        valorizado = await conn.fetchval(
            """SELECT COALESCE(SUM(cantidad_disponible * costo_unitario), 0)
               FROM prod_inventario_ingresos
               WHERE item_id = $1 AND cantidad_disponible > 0""",
            item_id
        )

        return {
            "item": row_to_dict(item),
            "movimientos": movimientos,
            "saldo_actual": float(item['stock_actual']),
            "valorizado": round(float(valorizado), 2),
        }

# ==================== REPORTE ITEM - ESTADOS (PIVOT) ====================

@router.get("/reportes/estados-item")
async def get_reporte_estados_item(
    search: str = None,
    marca_id: str = None,
    tipo_id: str = None,
    entalle_id: str = None,
    tela_id: str = None,
    hilo_especifico_id: str = None,
    prioridad: str = None,  # urgente|normal
    include_tienda: bool = False,
):
    """Reporte tipo Power BI: ITEM (marca+tipo+entalle+tela) + HILO, columnas por estado = COUNT(registros)."""

    # Nota: en algunos entornos este estado aparece como "Para Atraque" (en DB/código) o "Para Atanque" (en reportes).
    # Ambos se consolidan en la misma columna (para_atanque).
    estados_map = {
        "Para Corte": "para_corte",
        "Para Costura": "para_costura",
        "Para Atraque": "para_atanque",
        "Para Atanque": "para_atanque",
        "Para Lavandería": "para_lavanderia",
        "Acabado": "acabado",
        "Almacén PT": "almacen_pt",
        "Tienda": "tienda",
    }

    estados_incluidos = [
        "Para Corte",
        "Para Costura",
        "Para Atraque",
        "Para Atanque",
        "Para Lavandería",
        "Acabado",
        "Almacén PT",
    ]
    if include_tienda:
        estados_incluidos.append("Tienda")

    pool = await get_pool()
    async with pool.acquire() as conn:
        query = """
            SELECT r.estado, r.urgente, he.nombre as hilo_nombre,
                   ma.nombre as marca_nombre,
                   t.nombre as tipo_nombre,
                   e.nombre as entalle_nombre,
                   te.nombre as tela_nombre
            FROM prod_registros r
            LEFT JOIN prod_modelos m ON r.modelo_id = m.id
            LEFT JOIN prod_marcas ma ON m.marca_id = ma.id
            LEFT JOIN prod_tipos t ON m.tipo_id = t.id
            LEFT JOIN prod_entalles e ON m.entalle_id = e.id
            LEFT JOIN prod_telas te ON m.tela_id = te.id
            LEFT JOIN prod_hilos_especificos he ON r.hilo_especifico_id = he.id
            WHERE 1=1
        """
        params = []

        if marca_id:
            params.append(marca_id)
            query += f" AND m.marca_id = ${len(params)}"
        if tipo_id:
            params.append(tipo_id)
            query += f" AND m.tipo_id = ${len(params)}"
        if entalle_id:
            params.append(entalle_id)
            query += f" AND m.entalle_id = ${len(params)}"
        if tela_id:
            params.append(tela_id)
            query += f" AND m.tela_id = ${len(params)}"
        if hilo_especifico_id:
            params.append(hilo_especifico_id)
            query += f" AND r.hilo_especifico_id = ${len(params)}"

        if prioridad == "urgente":
            query += " AND r.urgente = true"
        elif prioridad == "normal":
            query += " AND (r.urgente = false OR r.urgente IS NULL)"

        rows = await conn.fetch(query, *params)

    data = {}

    def safe(x):
        return (x or '').strip()

    for row in rows:
        estado = row.get('estado')
        if estado not in estados_incluidos:
            continue

        marca = safe(row.get('marca_nombre')) or 'Sin Marca'
        tipo = safe(row.get('tipo_nombre')) or 'Sin Tipo'
        entalle = safe(row.get('entalle_nombre')) or 'Sin Entalle'
        tela = safe(row.get('tela_nombre')) or 'Sin Tela'
        hilo = safe(row.get('hilo_nombre')) or 'Sin Hilo'

        item = f"{marca} - {tipo} - {entalle} - {tela}"

        if search and search.strip().lower() not in item.lower():
            continue

        key = (item, hilo)
        if key not in data:
            data[key] = {
                "item": item,
                "hilo": hilo,
                "total": 0,
            }
            for est in estados_incluidos:
                data[key][estados_map[est]] = 0

        col = estados_map.get(estado)
        if not col:
            continue

        data[key][col] += 1
        data[key]["total"] += 1

    result_rows = list(data.values())
    result_rows.sort(key=lambda r: (r.get('item') or '', r.get('hilo') or ''))

    updated_at = datetime.now().strftime('%d/%m/%Y %H:%M')
    return {
        "updated_at": updated_at,
        "include_tienda": include_tienda,
        "rows": result_rows,
    }



@router.get("/reportes/estados-item/detalle")
async def get_reporte_estados_item_detalle(
    item: str,
    hilo: str,
    estado: str,
    include_tienda: bool = False,
    limit: int = 50,
    offset: int = 0,
    current_user: dict = Depends(get_current_user),
):
    """Detalle (drill-down) del reporte por Item+Hilo y Estado.

    - item: string exacto como se muestra ("Marca - Tipo - Entalle - Tela")
    - hilo: nombre del hilo específico (o "Sin Hilo")
    - estado: estado a filtrar (ej: "Para Costura")
    - include_tienda: si es False, no permite estado="Tienda"
    """

    if (not include_tienda) and estado == "Tienda":
        raise HTTPException(status_code=400, detail="Estado 'Tienda' no permitido cuando include_tienda=false")

    # Mapeo compatible: aceptar Para Atanque como sinónimo de Para Atraque
    if estado == "Para Atanque":
        estado = "Para Atraque"

    pool = await get_pool()
    async with pool.acquire() as conn:
        query = """
            SELECT r.id, r.n_corte, r.estado, r.urgente, r.fecha_creacion,
                   COALESCE(m.nombre, r.modelo_manual->>'nombre_modelo') as modelo_nombre,
                   COALESCE(ma.nombre, r.modelo_manual->>'marca_texto') as marca_nombre,
                   COALESCE(t.nombre, r.modelo_manual->>'tipo_texto') as tipo_nombre,
                   COALESCE(e.nombre, r.modelo_manual->>'entalle_texto') as entalle_nombre,
                   COALESCE(te.nombre, r.modelo_manual->>'tela_texto') as tela_nombre,
                   COALESCE(he.nombre, r.modelo_manual->>'hilo_especifico_texto') as hilo_nombre
            FROM prod_registros r
            LEFT JOIN prod_modelos m ON r.modelo_id = m.id
            LEFT JOIN prod_marcas ma ON m.marca_id = ma.id
            LEFT JOIN prod_tipos t ON m.tipo_id = t.id
            LEFT JOIN prod_entalles e ON m.entalle_id = e.id
            LEFT JOIN prod_telas te ON m.tela_id = te.id
            LEFT JOIN prod_hilos_especificos he ON r.hilo_especifico_id = he.id
            WHERE 1=1
        """
        params = []

        # Reconstruir item igual que el reporte
        params.append(item)
        query += f" AND (COALESCE(ma.nombre,'Sin Marca') || ' - ' || COALESCE(t.nombre,'Sin Tipo') || ' - ' || COALESCE(e.nombre,'Sin Entalle') || ' - ' || COALESCE(te.nombre,'Sin Tela')) = ${len(params)}"

        params.append(estado)
        query += f" AND r.estado = ${len(params)}"

        if hilo == "Sin Hilo":
            query += " AND (r.hilo_especifico_id IS NULL)"
        else:
            params.append(hilo)
            query += f" AND he.nombre = ${len(params)}"

        # Paginación
        params.append(limit)
        query += f" ORDER BY r.fecha_creacion DESC NULLS LAST LIMIT ${len(params)}"
        params.append(offset)
        query += f" OFFSET ${len(params)}"

        rows = await conn.fetch(query, *params)

        # total
        count_query = """
            SELECT COUNT(*)
            FROM prod_registros r
            LEFT JOIN prod_modelos m ON r.modelo_id = m.id
            LEFT JOIN prod_marcas ma ON m.marca_id = ma.id
            LEFT JOIN prod_tipos t ON m.tipo_id = t.id
            LEFT JOIN prod_entalles e ON m.entalle_id = e.id
            LEFT JOIN prod_telas te ON m.tela_id = te.id
            LEFT JOIN prod_hilos_especificos he ON r.hilo_especifico_id = he.id
            WHERE 1=1
        """
        count_params = []
        count_params.append(item)
        count_query += f" AND (COALESCE(ma.nombre,'Sin Marca') || ' - ' || COALESCE(t.nombre,'Sin Tipo') || ' - ' || COALESCE(e.nombre,'Sin Entalle') || ' - ' || COALESCE(te.nombre,'Sin Tela')) = ${len(count_params)}"
        count_params.append(estado)
        count_query += f" AND r.estado = ${len(count_params)}"
        if hilo == "Sin Hilo":
            count_query += " AND (r.hilo_especifico_id IS NULL)"
        else:
            count_params.append(hilo)
            count_query += f" AND he.nombre = ${len(count_params)}"

        total = await conn.fetchval(count_query, *count_params)

    result = []
    for r in rows:
        d = row_to_dict(r)
        # normalizar datetime
        if isinstance(d.get('fecha_creacion'), datetime):
            d['fecha_creacion'] = d['fecha_creacion'].strftime('%d/%m/%Y %H:%M')
        result.append(d)

    return {
        "item": item,
        "hilo": hilo,
        "estado": estado,
        "total": int(total or 0),
        "limit": limit,
        "offset": offset,
        "rows": result,
    }

@router.get("/reportes/estados-item/export")
async def export_reporte_estados_item(
    search: str = None,
    marca_id: str = None,
    tipo_id: str = None,
    entalle_id: str = None,
    tela_id: str = None,
    hilo_especifico_id: str = None,
    prioridad: str = None,
    include_tienda: bool = False,
    current_user: dict = Depends(get_current_user),
):
    """Export CSV (Excel) del reporte ITEM - ESTADOS."""

    reporte = await get_reporte_estados_item(
        search=search,
        marca_id=marca_id,
        tipo_id=tipo_id,
        entalle_id=entalle_id,
        tela_id=tela_id,
        hilo_especifico_id=hilo_especifico_id,
        prioridad=prioridad,
        include_tienda=include_tienda,
    )

    cols = [
        ('Item', 'item'),
        ('Hilo', 'hilo'),
        ('Para Corte', 'para_corte'),
        ('Para Costura', 'para_costura'),
        ('Para Atraque', 'para_atanque'),
        ('Para Lavandería', 'para_lavanderia'),
        ('Acabado', 'acabado'),
        ('Almacén PT', 'almacen_pt'),
    ]
    if include_tienda:
        cols.append(('Tienda', 'tienda'))
    cols.append(('Total', 'total'))

    output = io.StringIO()
    output.write('\ufeff')
    output.write(','.join([c[0] for c in cols]) + '\n')

    for row in reporte.get('rows', []):
        values = []
        for _, key in cols:
            val = row.get(key)
            if val is None:
                values.append('')
            else:
                s = str(val).replace('"', '""')
                if ',' in s or '"' in s or '\n' in s:
                    s = f'"{s}"'
                values.append(s)
        output.write(','.join(values) + '\n')

    output.seek(0)
    filename = f"reporte_estados_item_{datetime.now().strftime('%Y%m%d')}.csv"

    return StreamingResponse(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )

# ==================== ENDPOINTS BACKUP ====================

BACKUP_TABLES = [
    # Catálogos
    'prod_marcas', 'prod_tipos', 'prod_entalles', 'prod_telas', 'prod_hilos',
    'prod_hilos_especificos', 'prod_tallas_catalogo', 'prod_colores_generales',
    'prod_colores_catalogo',
    # Maestros
    'prod_bases', 'prod_modelos', 'prod_rutas_produccion',
    'prod_servicios_produccion', 'prod_personas_produccion',
    # Registros y producción
    'prod_registros', 'prod_registro_tallas', 'prod_movimientos_produccion',
    'prod_mermas', 'prod_fallados', 'prod_registro_arreglos', 'prod_guias_remision',
    # Inventario
    'prod_inventario', 'prod_inventario_ingresos', 'prod_inventario_salidas',
    'prod_inventario_ajustes', 'prod_inventario_rollos',
    # Requerimiento y reservas
    'prod_registro_requerimiento_mp', 'prod_inventario_reservas',
    'prod_inventario_reservas_linea',
    # Incidencias y paralizaciones
    'prod_incidencia', 'prod_motivos_incidencia', 'prod_paralizacion',
    # BOM
    'prod_bom_cabecera', 'prod_modelo_bom_linea', 'prod_modelo_tallas',
    # Configuración y sistema
    'prod_configuracion', 'prod_conversacion', 'prod_avance_historial',
    'prod_registro_cierre', 'prod_lineas_negocio',
    # Usuarios y empresas
    'prod_usuarios', 'prod_empresas',
]

@router.get("/backup/create")
async def create_backup(current_user: dict = Depends(get_current_user)):
    """Crea un backup completo de todas las tablas"""
    if current_user['rol'] != 'admin':
        raise HTTPException(status_code=403, detail="Solo administradores pueden crear backups")
    
    pool = await get_pool()
    backup_data = {
        "version": "1.0",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": current_user['username'],
        "tables": {}
    }
    
    async with pool.acquire() as conn:
        for table in BACKUP_TABLES:
            try:
                # Verificar que la tabla existe antes de consultar
                exists = await conn.fetchval(
                    "SELECT EXISTS(SELECT 1 FROM information_schema.tables WHERE table_name = $1)",
                    table
                )
                if not exists:
                    continue
                rows = await conn.fetch(f"SELECT * FROM {table}")
                table_data = []
                for row in rows:
                    row_dict = dict(row)
                    # Convertir tipos no serializables
                    for key, value in row_dict.items():
                        if isinstance(value, datetime):
                            row_dict[key] = value.isoformat() + "Z"
                        elif isinstance(value, date):
                            row_dict[key] = value.isoformat()
                        elif isinstance(value, uuid.UUID):
                            row_dict[key] = str(value)
                        elif isinstance(value, Decimal):
                            row_dict[key] = float(value)
                        elif isinstance(value, bytes):
                            row_dict[key] = value.decode('utf-8', errors='replace')
                        elif isinstance(value, (list, dict)):
                            pass  # JSON-compatible, keep as-is
                    table_data.append(row_dict)
                backup_data["tables"][table] = table_data
            except Exception as e:
                backup_data["tables"][table] = {"error": str(e), "rows": 0}
    
    # Registrar actividad
    await registrar_actividad(
        pool,
        usuario_id=current_user['id'],
        usuario_nombre=current_user['username'],
        tipo_accion="crear",
        tabla_afectada="backup",
        descripcion="Creó backup completo de la base de datos"
    )
    
    # Generar archivo JSON
    json_content = json.dumps(backup_data, ensure_ascii=False, indent=2)
    filename = f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    
    return StreamingResponse(
        io.BytesIO(json_content.encode('utf-8')),
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/backup/info")
async def backup_info(current_user: dict = Depends(get_current_user)):
    """Retorna información sobre las tablas para backup"""
    if current_user['rol'] != 'admin':
        raise HTTPException(status_code=403, detail="Solo administradores")
    
    pool = await get_pool()
    info = {"tables": []}
    
    async with pool.acquire() as conn:
        for table in BACKUP_TABLES:
            try:
                exists = await conn.fetchval(
                    "SELECT EXISTS(SELECT 1 FROM information_schema.tables WHERE table_name = $1)",
                    table
                )
                if not exists:
                    info["tables"].append({"name": table, "count": 0, "missing": True})
                    continue
                count = await conn.fetchval(f"SELECT COUNT(*) FROM {table}")
                info["tables"].append({"name": table, "count": count})
            except Exception:
                info["tables"].append({"name": table, "count": 0, "error": True})
    
    return info

@router.post("/backup/restore")
async def restore_backup(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    """Restaura un backup desde archivo JSON"""
    if current_user['rol'] != 'admin':
        raise HTTPException(status_code=403, detail="Solo administradores pueden restaurar backups")
    
    try:
        content = await file.read()
        backup_data = json.loads(content.decode('utf-8'))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error al leer archivo: {str(e)}")
    
    if "tables" not in backup_data:
        raise HTTPException(status_code=400, detail="Formato de backup inválido")
    
    pool = await get_pool()
    restored = []
    errors = []
    
    async with pool.acquire() as conn:
        for table, rows in backup_data["tables"].items():
            if table not in BACKUP_TABLES:
                continue
            if isinstance(rows, dict) and "error" in rows:
                continue
            if not rows:
                continue
            
            try:
                # Eliminar datos existentes
                await conn.execute(f"DELETE FROM {table}")
                
                # Insertar nuevos datos
                for row in rows:
                    columns = list(row.keys())
                    values = list(row.values())
                    placeholders = [f"${i+1}" for i in range(len(columns))]
                    
                    query = f"INSERT INTO {table} ({', '.join(columns)}) VALUES ({', '.join(placeholders)})"
                    try:
                        await conn.execute(query, *values)
                    except Exception as row_error:
                        errors.append(f"{table}: {str(row_error)[:50]}")
                
                restored.append(table)
            except Exception as table_error:
                errors.append(f"{table}: {str(table_error)}")
    
    # Registrar actividad
    await registrar_actividad(
        pool,
        usuario_id=current_user['id'],
        usuario_nombre=current_user['username'],
        tipo_accion="editar",
        tabla_afectada="backup",
        descripcion=f"Restauró backup: {len(restored)} tablas restauradas",
        datos_nuevos={"tablas_restauradas": restored, "errores": errors}
    )
    
    return {
        "message": "Backup restaurado",
        "restored_tables": restored,
        "errors": errors
    }

# ==================== ENDPOINTS EXPORTAR EXCEL ====================

EXPORT_TITLES = {
    "registros": "Registros de Producción",
    "inventario": "Inventario",
    "movimientos": "Movimientos de Inventario",
    "productividad": "Productividad",
    "personas": "Personas de Producción",
    "modelos": "Modelos",
    "mermas": "Mermas",
    "fallados": "Fallados",
    "arreglos": "Arreglos",
    "productos_odoo": "Productos Odoo",
    "wip": "Valorizado WIP en Proceso",
}

EXPORT_CONFIG = {
    "registros": {
        "query": """
            SELECT r.n_corte, r.fecha_creacion, r.estado, r.urgente,
                   m.nombre as modelo, ma.nombre as marca, t.nombre as tipo,
                   en.nombre as entalle, te.nombre as tela,
                   h.nombre as hilo, he.nombre as hilo_especifico,
                   r.curva
            FROM prod_registros r
            LEFT JOIN prod_modelos m ON r.modelo_id = m.id
            LEFT JOIN prod_marcas ma ON m.marca_id = ma.id
            LEFT JOIN prod_tipos t ON m.tipo_id = t.id
            LEFT JOIN prod_entalles en ON m.entalle_id = en.id
            LEFT JOIN prod_telas te ON m.tela_id = te.id
            LEFT JOIN prod_hilos h ON m.hilo_id = h.id
            LEFT JOIN prod_hilos_especificos he ON m.hilo_especifico_id = he.id
            ORDER BY r.fecha_creacion DESC
        """,
        "headers": ["N° Corte", "Fecha", "Estado", "Urgente", "Modelo", "Marca", "Tipo", "Entalle", "Tela", "Hilo", "Hilo Específico", "Curva"]
    },
    "inventario": {
        "query": """
            SELECT codigo, nombre, descripcion, unidad_medida,
                   stock_actual, stock_minimo, control_por_rollos,
                   COALESCE(costo_promedio, costo_compra, 0)::numeric(14,4) AS costo_unitario,
                   (COALESCE(stock_actual, 0)
                    * COALESCE(costo_promedio, costo_compra, 0))::numeric(14,2) AS valorizado
            FROM prod_inventario ORDER BY codigo
        """,
        "headers": ["Código", "Nombre", "Descripción", "Unidad",
                    "Stock Actual", "Stock Mínimo", "Control Rollos",
                    "Costo Unitario", "Valorizado"]
    },
    "wip": {
        # Valorizado WIP: registros (cortes) en proceso.
        #   Costo MP       = SUM(costo_total) de prod_inventario_salidas (FIFO)
        #   Costo Servicios = SUM(costo_calculado) de prod_movimientos_produccion
        # Los registros activos usan `modelo_manual` (JSONB) con marca/tipo/etc
        # ya pre-resueltos como texto, por eso usamos COALESCE con esos campos.
        # Excluye registros CERRADOS / ANULADOS.
        "query": """
            SELECT r.n_corte AS corte,
                   COALESCE(m.nombre,    r.modelo_manual->>'nombre_modelo', '—') AS modelo,
                   COALESCE(ma.nombre,   r.modelo_manual->>'marca_texto',   '—') AS marca,
                   COALESCE(t.nombre,    r.modelo_manual->>'tipo_texto',    '—') AS tipo,
                   COALESCE(en.nombre,   r.modelo_manual->>'entalle_texto', '—') AS entalle,
                   COALESCE(te.nombre,   r.modelo_manual->>'tela_texto',    '—') AS tela,
                   COALESCE(h.nombre,    r.modelo_manual->>'hilo_texto',    '—') AS hilo,
                   COALESCE(r.estado_op, r.estado, '—') AS estado,
                   r.fecha_creacion,
                   COALESCE((SELECT SUM(cantidad_real)
                              FROM prod_registro_tallas
                             WHERE registro_id = r.id), 0)::int AS prendas,
                   ROUND(COALESCE((SELECT SUM(costo_total)
                                    FROM prod_inventario_salidas
                                   WHERE registro_id = r.id), 0)::numeric, 2) AS costo_mp,
                   ROUND(COALESCE((SELECT SUM(costo_calculado)
                                    FROM prod_movimientos_produccion
                                   WHERE registro_id = r.id), 0)::numeric, 2) AS costo_servicios,
                   ROUND((COALESCE((SELECT SUM(costo_total)
                                     FROM prod_inventario_salidas
                                    WHERE registro_id = r.id), 0)
                         + COALESCE((SELECT SUM(costo_calculado)
                                      FROM prod_movimientos_produccion
                                     WHERE registro_id = r.id), 0))::numeric, 2) AS total
              FROM prod_registros r
              LEFT JOIN prod_modelos m  ON r.modelo_id = m.id
              LEFT JOIN prod_marcas  ma ON m.marca_id   = ma.id
              LEFT JOIN prod_tipos   t  ON m.tipo_id    = t.id
              LEFT JOIN prod_entalles en ON m.entalle_id = en.id
              LEFT JOIN prod_telas   te ON m.tela_id    = te.id
              LEFT JOIN prod_hilos   h  ON m.hilo_id    = h.id
             WHERE COALESCE(r.estado, '') NOT IN ('CERRADA', 'ANULADA')
               AND COALESCE(r.estado_op, '') NOT IN ('CERRADO', 'ENTREGADO')
             ORDER BY r.fecha_creacion DESC NULLS LAST, r.n_corte
        """,
        "headers": ["N° Corte", "Modelo", "Marca", "Tipo", "Entalle", "Tela", "Hilo",
                    "Estado", "Fecha", "Prendas",
                    "Costo MP", "Costo Servicios", "Total"]
    },
    "productos_odoo": {
        # NOTAS sobre IGV:
        #  - list_price en Odoo viene con IGV incluido (precio público)
        #  - costo_manual es neto (sin IGV)
        #  - Para calcular margen REAL: precio_neto = list_price / 1.18,
        #    margen = precio_neto - costo
        "query": """
            SELECT pe.odoo_default_code AS sku,
                   pe.odoo_nombre        AS producto,
                   COALESCE(ma.nombre, pe.odoo_marca_texto, '—')   AS marca,
                   COALESCE(ti.nombre, pe.odoo_tipo_texto, '—')    AS tipo,
                   COALESCE(en.nombre, pe.odoo_entalle_texto, '—') AS entalle,
                   COALESCE(tg.nombre, te.nombre, pe.odoo_tela_texto, '—') AS tela,
                   COALESCE(pe.costo_manual, 0)::numeric(14,2)     AS costo,
                   COALESCE(pt.list_price, 0)::numeric(14,2)       AS precio_venta_con_igv,
                   ROUND((COALESCE(pt.list_price, 0) / 1.18)::numeric, 2) AS precio_venta_sin_igv,
                   ROUND(((COALESCE(pt.list_price, 0) / 1.18) - COALESCE(pe.costo_manual, 0))::numeric, 2)
                                                                   AS margen_real,
                   CASE WHEN COALESCE(pt.list_price, 0) > 0
                        THEN ROUND(
                              (((COALESCE(pt.list_price, 0) / 1.18) - COALESCE(pe.costo_manual, 0))
                               / (pt.list_price / 1.18) * 100)::numeric, 2)
                        ELSE 0
                   END AS margen_pct,
                   COALESCE(pe.odoo_stock_actual, 0)::numeric(14,2) AS stock,
                   (COALESCE(pe.odoo_stock_actual, 0) * COALESCE(pe.costo_manual, 0))::numeric(14,2)
                                                                   AS valorizado_costo,
                   ROUND((COALESCE(pe.odoo_stock_actual, 0) * COALESCE(pt.list_price, 0) / 1.18)::numeric, 2)
                                                                   AS valorizado_venta_neto,
                   pe.estado AS estado_clasificacion
              FROM prod_odoo_productos_enriq pe
              LEFT JOIN odoo.product_template pt ON pt.odoo_id = pe.odoo_template_id
              LEFT JOIN prod_marcas    ma ON ma.id = pe.marca_id
              LEFT JOIN prod_tipos     ti ON ti.id = pe.tipo_id
              LEFT JOIN prod_entalles  en ON en.id = pe.entalle_id
              LEFT JOIN prod_telas     te ON te.id = pe.tela_id
              LEFT JOIN prod_telas_general tg ON tg.id = pe.tela_general_id
             ORDER BY pe.odoo_nombre
        """,
        "headers": ["SKU", "Producto", "Marca", "Tipo", "Entalle", "Tela",
                    "Costo", "Precio Venta c/IGV", "Precio Venta s/IGV",
                    "Margen S/", "Margen %",
                    "Stock", "Valorizado Costo", "Valorizado Venta s/IGV",
                    "Estado"]
    },
    "movimientos": {
        "query": """
            SELECT i.codigo, i.nombre,
                   COALESCE(ing.fecha, sal.fecha, aj.fecha) as fecha,
                   CASE
                       WHEN ing.id IS NOT NULL THEN 'Ingreso'
                       WHEN sal.id IS NOT NULL THEN 'Salida'
                       WHEN aj.id IS NOT NULL THEN 'Ajuste'
                   END as tipo,
                   COALESCE(ing.cantidad, -sal.cantidad, aj.cantidad) as cantidad,
                   COALESCE(ing.costo_unitario, 0) as costo
            FROM prod_inventario i
            LEFT JOIN prod_inventario_ingresos ing ON i.id = ing.inventario_id
            LEFT JOIN prod_inventario_salidas sal ON i.id = sal.inventario_id
            LEFT JOIN prod_inventario_ajustes aj ON i.id = aj.inventario_id
            WHERE ing.id IS NOT NULL OR sal.id IS NOT NULL OR aj.id IS NOT NULL
            ORDER BY COALESCE(ing.fecha, sal.fecha, aj.fecha) DESC
        """,
        "headers": ["Código", "Item", "Fecha", "Tipo", "Cantidad", "Costo"]
    },
    "productividad": {
        "query": """
            SELECT p.nombre as persona, s.nombre as servicio,
                   mp.cantidad_enviada as cantidad, mp.costo_calculado as monto,
                   mp.fecha_inicio as fecha, r.n_corte, mp.observaciones
            FROM prod_movimientos_produccion mp
            LEFT JOIN prod_personas_produccion p ON mp.persona_id = p.id
            LEFT JOIN prod_servicios_produccion s ON mp.servicio_id = s.id
            LEFT JOIN prod_registros r ON mp.registro_id = r.id
            ORDER BY mp.created_at DESC
        """,
        "headers": ["Persona", "Servicio", "Cantidad", "Monto", "Fecha", "N° Corte", "Observaciones"]
    },
    "personas": {
        "query": "SELECT nombre, telefono, activo FROM prod_personas_produccion ORDER BY nombre",
        "headers": ["Nombre", "Teléfono", "Activo"]
    },
    "modelos": {
        "query": """
            SELECT m.nombre, ma.nombre as marca, t.nombre as tipo,
                   e.nombre as entalle, te.nombre as tela
            FROM prod_modelos m
            LEFT JOIN prod_marcas ma ON m.marca_id = ma.id
            LEFT JOIN prod_tipos t ON m.tipo_id = t.id
            LEFT JOIN prod_entalles e ON m.entalle_id = e.id
            LEFT JOIN prod_telas te ON m.tela_id = te.id
            ORDER BY m.nombre
        """,
        "headers": ["Nombre", "Marca", "Tipo", "Entalle", "Tela"]
    },
    "mermas": {
        "query": """
            SELECT r.n_corte, sp.nombre as servicio, pp.nombre as persona,
                   m.cantidad, m.tipo, m.motivo, m.fecha
            FROM prod_mermas m
            LEFT JOIN prod_registros r ON m.registro_id = r.id
            LEFT JOIN prod_servicios_produccion sp ON m.servicio_id = sp.id
            LEFT JOIN prod_personas_produccion pp ON m.persona_id = pp.id
            ORDER BY m.fecha DESC NULLS LAST
        """,
        "headers": ["N° Corte", "Servicio", "Persona", "Cantidad", "Tipo", "Motivo", "Fecha"]
    },
    "fallados": {
        "query": """
            SELECT r.n_corte, sp.nombre as servicio_deteccion,
                   f.cantidad_detectada, f.cantidad_reparable, f.cantidad_no_reparable,
                   f.motivo, f.estado, f.destino_no_reparable, f.fecha_deteccion
            FROM prod_fallados f
            LEFT JOIN prod_registros r ON f.registro_id = r.id
            LEFT JOIN prod_servicios_produccion sp ON f.servicio_detectado_id = sp.id
            ORDER BY f.created_at DESC
        """,
        "headers": ["N° Corte", "Servicio Detección", "Detectadas", "Reparables", "No Reparables", "Motivo", "Estado", "Destino", "Fecha"]
    },
    "arreglos": {
        "query": """
            SELECT r.n_corte, sp.nombre as servicio, pp.nombre as persona,
                   a.cantidad, a.cantidad_recuperada, a.cantidad_liquidacion, a.cantidad_merma,
                   a.estado, a.fecha_envio, a.fecha_limite
            FROM prod_registro_arreglos a
            LEFT JOIN prod_registros r ON a.registro_id = r.id
            LEFT JOIN prod_servicios_produccion sp ON a.servicio_id = sp.id
            LEFT JOIN prod_personas_produccion pp ON a.persona_id = pp.id
            ORDER BY a.created_at DESC
        """,
        "headers": ["N° Corte", "Servicio", "Persona", "Cantidad", "Recuperado", "Liquidación", "Merma", "Estado", "F. Envío", "F. Límite"]
    }
}

def _format_cell(val):
    """Formatea un valor para exportación."""
    if val is None:
        return ""
    if isinstance(val, bool):
        return "Sí" if val else "No"
    if isinstance(val, (datetime, date)):
        return val.strftime("%d/%m/%Y")
    return str(val)

def _get_export_data(rows, headers):
    """Convierte filas de DB en lista de dicts formateados."""
    data = []
    for row in rows:
        data.append([_format_cell(v) for v in row.values()])
    return data


def _build_filtered_query(tabla: str, params: dict):
    """Construye query con filtros dinámicos para exportación."""
    config = EXPORT_CONFIG[tabla]
    query = config["query"]

    if tabla == "registros":
        conditions = []
        args = []
        arg_idx = 1

        if params.get("marca_id"):
            ids = [int(x) for x in params["marca_id"].split(",") if x.strip()]
            if ids:
                placeholders = ", ".join(f"${arg_idx + i}" for i in range(len(ids)))
                conditions.append(f"ma.id IN ({placeholders})")
                args.extend(ids)
                arg_idx += len(ids)

        if params.get("tipo_id"):
            ids = [int(x) for x in params["tipo_id"].split(",") if x.strip()]
            if ids:
                placeholders = ", ".join(f"${arg_idx + i}" for i in range(len(ids)))
                conditions.append(f"t.id IN ({placeholders})")
                args.extend(ids)
                arg_idx += len(ids)

        if params.get("entalle_id"):
            ids = [int(x) for x in params["entalle_id"].split(",") if x.strip()]
            if ids:
                placeholders = ", ".join(f"${arg_idx + i}" for i in range(len(ids)))
                conditions.append(f"en.id IN ({placeholders})")
                args.extend(ids)
                arg_idx += len(ids)

        if params.get("tela_id"):
            ids = [int(x) for x in params["tela_id"].split(",") if x.strip()]
            if ids:
                placeholders = ", ".join(f"${arg_idx + i}" for i in range(len(ids)))
                conditions.append(f"te.id IN ({placeholders})")
                args.extend(ids)
                arg_idx += len(ids)

        if params.get("estados"):
            estados = [x.strip() for x in params["estados"].split(",") if x.strip()]
            if estados:
                placeholders = ", ".join(f"${arg_idx + i}" for i in range(len(estados)))
                conditions.append(f"r.estado IN ({placeholders})")
                args.extend(estados)
                arg_idx += len(estados)

        if params.get("excluir_estados"):
            estados = [x.strip() for x in params["excluir_estados"].split(",") if x.strip()]
            if estados:
                placeholders = ", ".join(f"${arg_idx + i}" for i in range(len(estados)))
                conditions.append(f"r.estado NOT IN ({placeholders})")
                args.extend(estados)
                arg_idx += len(estados)

        if params.get("search"):
            conditions.append(f"(r.n_corte ILIKE ${arg_idx} OR m.nombre ILIKE ${arg_idx})")
            args.append(f"%{params['search']}%")
            arg_idx += 1

        if conditions:
            # Insert WHERE before ORDER BY
            order_idx = query.upper().rfind("ORDER BY")
            if order_idx > 0:
                query = query[:order_idx] + " WHERE " + " AND ".join(conditions) + " " + query[order_idx:]
            else:
                query += " WHERE " + " AND ".join(conditions)

        return query, args

    return query, []


async def _fetch_export_rows(tabla: str, params: dict):
    """Fetch rows con filtros aplicados."""
    query, args = _build_filtered_query(tabla, params)
    pool = await get_pool()
    async with pool.acquire() as conn:
        return await conn.fetch(query, *args)


@router.get("/export/{tabla}")
async def export_tabla(
    tabla: str,
    request: Request,
    format: str = "xlsx",
    current_user: dict = Depends(get_current_user),
):
    """Exporta una tabla a formato Excel (.xlsx) o PDF con filtros."""
    if tabla not in EXPORT_CONFIG:
        raise HTTPException(status_code=400, detail=f"Tabla '{tabla}' no exportable")

    # Extraer filtros del query string (soporta marca_id/marca_ids, etc.)
    filter_params = {}
    filter_keys = {
        "marca_id": ["marca_id", "marca_ids"],
        "tipo_id": ["tipo_id", "tipo_ids"],
        "entalle_id": ["entalle_id", "entalle_ids"],
        "tela_id": ["tela_id", "tela_ids"],
        "estados": ["estados"],
        "excluir_estados": ["excluir_estados"],
        "search": ["search"],
    }
    for canonical, aliases in filter_keys.items():
        for alias in aliases:
            val = request.query_params.get(alias)
            if val:
                filter_params[canonical] = val
                break

    if format == "pdf":
        return await _export_to_pdf(tabla, filter_params)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    config = EXPORT_CONFIG[tabla]
    titulo = EXPORT_TITLES.get(tabla, tabla.title())

    rows = await _fetch_export_rows(tabla, filter_params)

    data = _get_export_data(rows, config["headers"])

    wb = Workbook()
    ws = wb.active
    ws.title = titulo[:31]

    # Estilos
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    title_font = Font(name="Calibri", bold=True, size=14, color="1E3A5F")
    subtitle_font = Font(name="Calibri", size=10, color="6B7280")
    cell_font = Font(name="Calibri", size=10)
    cell_border = Border(
        bottom=Side(style="thin", color="E5E7EB"),
    )
    alt_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

    # Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(config["headers"]))
    title_cell = ws.cell(row=1, column=1, value=titulo)
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    # Subtítulo con fecha
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(config["headers"]))
    sub_cell = ws.cell(row=2, column=1, value=f"Exportado el {datetime.now().strftime('%d/%m/%Y %H:%M')}  —  {len(data)} registros")
    sub_cell.font = subtitle_font
    ws.row_dimensions[2].height = 20

    # Headers (fila 4)
    header_row = 4
    for col_idx, header in enumerate(config["headers"], 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = Border(
            bottom=Side(style="medium", color="1E40AF"),
        )
    ws.row_dimensions[header_row].height = 28

    # Datos
    for row_idx, row_data in enumerate(data, header_row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.border = cell_border
            cell.alignment = Alignment(vertical="center")
            # Filas alternas
            if (row_idx - header_row) % 2 == 0:
                cell.fill = alt_fill

    # Auto-ajustar anchos
    for col_idx, header in enumerate(config["headers"], 1):
        max_len = len(header)
        for row_data in data[:100]:  # Solo primeras 100 filas para performance
            if col_idx - 1 < len(row_data):
                max_len = max(max_len, len(str(row_data[col_idx - 1])))
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # Congelar headers
    ws.freeze_panes = f"A{header_row + 1}"

    # Auto-filtro
    last_col_letter = get_column_letter(len(config["headers"]))
    ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{header_row + len(data)}"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{tabla}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


async def _export_to_pdf(tabla: str, filter_params: dict = None):
    """Exporta una tabla a PDF con formato profesional."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer

    if tabla not in EXPORT_CONFIG:
        raise HTTPException(status_code=400, detail=f"Tabla '{tabla}' no exportable")

    config = EXPORT_CONFIG[tabla]
    titulo = EXPORT_TITLES.get(tabla, tabla.title())

    rows = await _fetch_export_rows(tabla, filter_params or {})

    data = _get_export_data(rows, config["headers"])

    output = io.BytesIO()
    page_size = landscape(A4) if len(config["headers"]) > 7 else A4
    doc = SimpleDocTemplate(output, pagesize=page_size, topMargin=20*mm, bottomMargin=15*mm, leftMargin=12*mm, rightMargin=12*mm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("ExportTitle", parent=styles["Heading1"], fontSize=16, textColor=colors.HexColor("#1E3A5F"), spaceAfter=4)
    subtitle_style = ParagraphStyle("ExportSubtitle", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#6B7280"), spaceAfter=12)
    cell_style = ParagraphStyle("CellStyle", parent=styles["Normal"], fontSize=7, leading=9)

    elements = []
    elements.append(Paragraph(titulo, title_style))
    elements.append(Paragraph(f"Exportado el {datetime.now().strftime('%d/%m/%Y %H:%M')}  —  {len(data)} registros", subtitle_style))

    # Preparar datos de tabla con word-wrap via Paragraph
    table_data = [config["headers"]]
    for row_data in data:
        table_data.append([Paragraph(str(v), cell_style) if len(str(v)) > 15 else str(v) for v in row_data])

    # Calcular anchos proporcionales
    avail_width = page_size[0] - 24*mm
    n_cols = len(config["headers"])
    col_widths = [avail_width / n_cols] * n_cols

    t = RLTable(table_data, colWidths=col_widths, repeatRows=1)

    header_bg = colors.HexColor("#2563EB")
    alt_bg = colors.HexColor("#F9FAFB")
    border_color = colors.HexColor("#E5E7EB")

    style_cmds = [
        # Header
        ("BACKGROUND", (0, 0), (-1, 0), header_bg),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, 0), 6),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        # Body
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("TOPPADDING", (0, 1), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        # Grid
        ("LINEBELOW", (0, 0), (-1, 0), 1, colors.HexColor("#1E40AF")),
        ("LINEBELOW", (0, 1), (-1, -1), 0.5, border_color),
    ]

    # Filas alternas
    for i in range(1, len(table_data)):
        if i % 2 == 0:
            style_cmds.append(("BACKGROUND", (0, i), (-1, i), alt_bg))

    t.setStyle(TableStyle(style_cmds))
    elements.append(t)

    doc.build(elements)
    output.seek(0)

    filename = f"{tabla}_{datetime.now().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ==================== REPORTE PARALIZADOS / DIVISIÓN DE LOTE ====================




# ==================== REPORTE PARALIZADOS ====================

@router.get("/reportes/paralizados")
async def reporte_paralizados(
    solo_activas: Optional[str] = None,
):
    """Reporte completo de paralizaciones: activas + historial."""
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch("""
            SELECT 
                p.id,
                p.registro_id,
                p.movimiento_id,
                p.fecha_inicio,
                p.fecha_fin,
                p.motivo,
                p.comentario,
                p.activa,
                r.n_corte,
                r.estado as registro_estado,
                r.urgente,
                COALESCE(mod.nombre, r.modelo_manual->>'nombre_modelo') as modelo_nombre,
                COALESCE(ma.nombre, r.modelo_manual->>'marca_texto') as marca_nombre,
                -- Último movimiento activo
                (SELECT sp.nombre FROM prod_movimientos_produccion mp
                 JOIN prod_servicios_produccion sp ON sp.id = mp.servicio_id
                 WHERE mp.registro_id = r.id AND mp.avance_porcentaje < 100
                 ORDER BY mp.fecha_inicio DESC NULLS LAST LIMIT 1) as servicio_actual,
                (SELECT pp.nombre FROM prod_movimientos_produccion mp
                 JOIN prod_personas_produccion pp ON pp.id = mp.persona_id
                 WHERE mp.registro_id = r.id AND mp.avance_porcentaje < 100
                 ORDER BY mp.fecha_inicio DESC NULLS LAST LIMIT 1) as persona_actual,
                -- Movimiento vinculado (si tiene)
                (SELECT sp.nombre FROM prod_servicios_produccion sp
                 JOIN prod_movimientos_produccion mp ON mp.id = p.movimiento_id AND sp.id = mp.servicio_id
                 ) as servicio_movimiento,
                -- Incidencia vinculada
                (SELECT i.tipo FROM prod_incidencia i 
                 WHERE i.paralizacion_id = p.id LIMIT 1) as incidencia_tipo,
                (SELECT i.estado FROM prod_incidencia i 
                 WHERE i.paralizacion_id = p.id LIMIT 1) as incidencia_estado,
                -- Cantidad de prendas
                COALESCE((SELECT SUM(rt.cantidad_real) FROM prod_registro_tallas rt 
                          WHERE rt.registro_id = r.id), 0) as prendas
            FROM prod_paralizacion p
            JOIN prod_registros r ON r.id = p.registro_id
            LEFT JOIN prod_modelos mod ON mod.id = r.modelo_id
            LEFT JOIN prod_marcas ma ON ma.id = mod.marca_id
            ORDER BY p.activa DESC, p.fecha_inicio DESC
        """)

        from datetime import date
        hoy = date.today()
        resultado = []
        resumen = {"activas": 0, "resueltas": 0, "total": 0, "prendas_afectadas": 0, "dias_promedio": 0}
        dias_list = []

        for row in rows:
            d = row_to_dict(row)
            activa = d.get("activa", False)

            # Filtro
            if solo_activas == "true" and not activa:
                continue

            # Calcular días — fi y ff pueden ser datetime, date o str (row_to_dict los pasa a ISO).
            def _to_date(x):
                if x is None:
                    return None
                if hasattr(x, 'date'):  # datetime
                    return x.date()
                if isinstance(x, str):
                    try:
                        # Quita sufijo Z si existe y toma solo la parte de fecha
                        return datetime.fromisoformat(x.replace('Z', '').split('T')[0]).date()
                    except Exception:
                        return None
                return x  # ya es date

            fi = d.get("fecha_inicio")
            ff = d.get("fecha_fin")
            fi_date = _to_date(fi)
            ff_date = _to_date(ff)
            if fi_date:
                if activa:
                    dias = (hoy - fi_date).days
                else:
                    dias = ((ff_date or hoy) - fi_date).days
            else:
                dias = 0

            dias_list.append(dias)

            if activa:
                resumen["activas"] += 1
                resumen["prendas_afectadas"] += int(d.get("prendas", 0) or 0)
            else:
                resumen["resueltas"] += 1
            resumen["total"] += 1

            # Serializar fechas
            for f in ("fecha_inicio", "fecha_fin"):
                if d.get(f):
                    d[f] = str(d[f])

            d["dias"] = dias
            d["servicio"] = d.get("servicio_movimiento") or d.get("servicio_actual") or ""
            d["persona"] = d.get("persona_actual") or ""
            resultado.append(d)

        if dias_list:
            resumen["dias_promedio"] = round(sum(dias_list) / len(dias_list), 1)

        # Motivos agrupados
        motivos_count = {}
        for r in resultado:
            m = r.get("motivo") or "Sin motivo"
            motivos_count[m] = motivos_count.get(m, 0) + 1
        motivos = [{"motivo": k, "cantidad": v} for k, v in sorted(motivos_count.items(), key=lambda x: -x[1])]

        return {
            "paralizaciones": resultado,
            "resumen": resumen,
            "motivos": motivos,
        }
