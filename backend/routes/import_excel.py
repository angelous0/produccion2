"""
Importación masiva de registros de producción desde Excel.
- GET  /api/registros/import-template  → descarga plantilla .xlsx
- POST /api/registros/import-validate  → valida sin guardar
- POST /api/registros/import-execute   → ejecuta importación
"""

from fastapi import APIRouter, UploadFile, File, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse
from datetime import datetime, date
import uuid, json, io, re

router = APIRouter(prefix="/api")

from db import get_pool
from auth_utils import get_current_user
from models import ESTADOS_PRODUCCION


# ─────────────────── helpers ───────────────────

TALLA_COLS = ["XS", "S", "M", "L", "XL", "XXL", "XXXL", "28", "30", "32", "34", "36", "38", "40"]

def _parse_date(val):
    """Parse dd/mm/yyyy o yyyy-mm-dd o datetime."""
    if val is None:
        return None
    if isinstance(val, (date, datetime)):
        return val if isinstance(val, date) and not isinstance(val, datetime) else val.date() if isinstance(val, datetime) else val
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Fecha inválida: {s}")


def _parse_num(val, default=0):
    if val is None:
        return default
    if isinstance(val, (int, float)):
        return val
    s = str(val).strip()
    if not s:
        return default
    return float(s)


def _clean_str(val):
    if val is None:
        return ""
    return str(val).strip()


def _read_excel(content: bytes):
    """Lee las 3 pestañas del Excel y devuelve listas de dicts."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

    def sheet_to_dicts(sheet):
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) < 2:
            return []
        headers = [str(h).strip() if h else "" for h in rows[0]]
        result = []
        for i, row in enumerate(rows[1:], start=2):
            d = {}
            for j, h in enumerate(headers):
                d[h] = row[j] if j < len(row) else None
            d["_row"] = i
            result.append(d)
        return result

    registros_data = sheet_to_dicts(wb["Registros"]) if "Registros" in wb.sheetnames else []
    movimientos_data = sheet_to_dicts(wb["Movimientos"]) if "Movimientos" in wb.sheetnames else []
    tallas_data = sheet_to_dicts(wb["Tallas"]) if "Tallas" in wb.sheetnames else []
    materiales_data = sheet_to_dicts(wb["Materiales"]) if "Materiales" in wb.sheetnames else []

    return registros_data, movimientos_data, tallas_data, materiales_data


async def _validate(registros_data, movimientos_data, tallas_data, materiales_data, conn):
    """Valida datos y devuelve errores/advertencias + datos procesados."""
    errors = []
    warnings = []

    # ── cargar catálogos ──
    servicios = {r["nombre"].strip().lower(): str(r["id"]) for r in await conn.fetch("SELECT id, nombre FROM prod_servicios_produccion")}
    personas = {r["nombre"].strip().lower(): str(r["id"]) for r in await conn.fetch("SELECT id, nombre FROM prod_personas_produccion")}
    lineas = {r["nombre"].strip().lower(): int(r["id"]) for r in await conn.fetch("SELECT id, nombre FROM finanzas2.cont_linea_negocio WHERE activo = true")}
    marcas = {r["nombre"].strip().lower(): str(r["id"]) for r in await conn.fetch("SELECT id, nombre FROM prod_marcas")}
    tipos = {r["nombre"].strip().lower(): str(r["id"]) for r in await conn.fetch("SELECT id, nombre FROM prod_tipos")}
    telas = {r["nombre"].strip().lower(): str(r["id"]) for r in await conn.fetch("SELECT id, nombre FROM prod_telas")}
    entalles = {r["nombre"].strip().lower(): str(r["id"]) for r in await conn.fetch("SELECT id, nombre FROM prod_entalles")}
    hilos = {r["nombre"].strip().lower(): str(r["id"]) for r in await conn.fetch("SELECT id, nombre FROM prod_hilos")}
    hilos_esp = {r["nombre"].strip().lower(): str(r["id"]) for r in await conn.fetch("SELECT id, nombre FROM prod_hilos_especificos")}
    tallas_cat = {r["nombre"].strip(): str(r["id"]) for r in await conn.fetch("SELECT id, nombre FROM prod_tallas_catalogo ORDER BY orden")}
    existing_cortes = {r["n_corte"] for r in await conn.fetch("SELECT n_corte FROM prod_registros")}

    # ── validar registros ──
    n_corte_map = {}  # n_corte → processed registro dict
    seen_cortes = set()

    for row in registros_data:
        rn = row["_row"]
        n_corte = _clean_str(row.get("N_Corte"))
        if not n_corte:
            errors.append({"row": rn, "sheet": "Registros", "msg": "N_Corte es obligatorio"})
            continue
        if n_corte in seen_cortes:
            errors.append({"row": rn, "sheet": "Registros", "msg": f"N_Corte '{n_corte}' duplicado en el Excel"})
            continue
        if n_corte in existing_cortes:
            errors.append({"row": rn, "sheet": "Registros", "msg": f"N_Corte '{n_corte}' ya existe en la base de datos"})
            continue
        seen_cortes.add(n_corte)

        estado = _clean_str(row.get("Estado_Actual")) or "Para Corte"
        if estado not in ESTADOS_PRODUCCION:
            errors.append({"row": rn, "sheet": "Registros", "msg": f"Estado '{estado}' no válido. Válidos: {', '.join(ESTADOS_PRODUCCION)}"})
            continue

        linea_nombre = _clean_str(row.get("Linea_Negocio")).lower()
        linea_id = lineas.get(linea_nombre) if linea_nombre else None
        if linea_nombre and not linea_id:
            errors.append({"row": rn, "sheet": "Registros", "msg": f"Línea de Negocio '{row.get('Linea_Negocio')}' no existe en catálogo"})
            continue

        try:
            fecha_inicio = _parse_date(row.get("Fecha_Inicio"))
        except ValueError as e:
            errors.append({"row": rn, "sheet": "Registros", "msg": str(e)})
            continue
        try:
            fecha_entrega = _parse_date(row.get("Fecha_Entrega"))
        except ValueError as e:
            errors.append({"row": rn, "sheet": "Registros", "msg": str(e)})
            continue

        prendas = int(_parse_num(row.get("Prendas_Total"), 0))
        urgente_str = _clean_str(row.get("Urgente")).upper()
        urgente = urgente_str in ("SI", "SÍ", "YES", "TRUE", "1")

        # Resolver catálogos para modelo_manual
        marca_txt = _clean_str(row.get("Marca"))
        tipo_txt = _clean_str(row.get("Tipo"))
        tela_txt = _clean_str(row.get("Tela"))
        entalle_txt = _clean_str(row.get("Entalle"))
        hilo_txt = _clean_str(row.get("Hilo"))
        hilo_esp_txt = _clean_str(row.get("Hilo_Especifico"))

        modelo_manual = {
            "nombre_modelo": _clean_str(row.get("Nombre_Modelo")) or None,
            "marca_id": marcas.get(marca_txt.lower()) if marca_txt else None,
            "marca_texto": marca_txt or None,
            "tipo_id": tipos.get(tipo_txt.lower()) if tipo_txt else None,
            "tipo_texto": tipo_txt or None,
            "tela_id": telas.get(tela_txt.lower()) if tela_txt else None,
            "tela_texto": tela_txt or None,
            "entalle_id": entalles.get(entalle_txt.lower()) if entalle_txt else None,
            "entalle_texto": entalle_txt or None,
            "hilo_id": hilos.get(hilo_txt.lower()) if hilo_txt else None,
            "hilo_texto": hilo_txt or None,
            "hilo_especifico_id": hilos_esp.get(hilo_esp_txt.lower()) if hilo_esp_txt else None,
            "hilo_especifico_texto": hilo_esp_txt or None,
        }

        if not _clean_str(row.get("Nombre_Modelo")):
            warnings.append({"row": rn, "sheet": "Registros", "msg": "Nombre_Modelo vacío"})

        reg_id = str(uuid.uuid4())
        n_corte_map[n_corte] = {
            "id": reg_id,
            "n_corte": n_corte,
            "estado": estado,
            "urgente": urgente,
            "linea_negocio_id": linea_id,
            "fecha_inicio": fecha_inicio,
            "fecha_entrega": fecha_entrega,
            "prendas_total": prendas,
            "observaciones": _clean_str(row.get("Observaciones")) or None,
            "modelo_manual": modelo_manual,
        }

    # ── validar tallas ──
    tallas_por_corte = {}  # n_corte → [{talla_id, nombre, cantidad}]
    for row in tallas_data:
        rn = row["_row"]
        n_corte = _clean_str(row.get("N_Corte"))
        if not n_corte:
            continue
        if n_corte not in n_corte_map:
            errors.append({"row": rn, "sheet": "Tallas", "msg": f"N_Corte '{n_corte}' no existe en pestaña Registros"})
            continue

        tallas_list = []
        for col in TALLA_COLS:
            val = row.get(col)
            cant = int(_parse_num(val, 0))
            if cant > 0:
                tid = tallas_cat.get(col)
                if not tid:
                    errors.append({"row": rn, "sheet": "Tallas", "msg": f"Talla '{col}' no existe en catálogo"})
                    break
                tallas_list.append({"talla_id": tid, "talla_nombre": col, "cantidad": cant})
        else:
            if tallas_list:
                tallas_por_corte[n_corte] = tallas_list

    # ── validar movimientos ──
    movimientos_por_corte = {}  # n_corte → [mov_dict]
    for row in movimientos_data:
        rn = row["_row"]
        n_corte = _clean_str(row.get("N_Corte"))
        if not n_corte:
            errors.append({"row": rn, "sheet": "Movimientos", "msg": "N_Corte es obligatorio"})
            continue
        if n_corte not in n_corte_map:
            errors.append({"row": rn, "sheet": "Movimientos", "msg": f"N_Corte '{n_corte}' no existe en pestaña Registros"})
            continue

        servicio_nombre = _clean_str(row.get("Servicio")).lower()
        if not servicio_nombre:
            errors.append({"row": rn, "sheet": "Movimientos", "msg": "Servicio es obligatorio"})
            continue
        servicio_id = servicios.get(servicio_nombre)
        if not servicio_id:
            errors.append({"row": rn, "sheet": "Movimientos", "msg": f"Servicio '{row.get('Servicio')}' no existe en catálogo"})
            continue

        persona_nombre = _clean_str(row.get("Persona")).lower()
        if not persona_nombre:
            errors.append({"row": rn, "sheet": "Movimientos", "msg": "Persona es obligatoria"})
            continue
        persona_id = personas.get(persona_nombre)
        if not persona_id:
            errors.append({"row": rn, "sheet": "Movimientos", "msg": f"Persona '{row.get('Persona')}' no existe en catálogo"})
            continue

        try:
            fi = _parse_date(row.get("Fecha_Inicio"))
        except ValueError as e:
            errors.append({"row": rn, "sheet": "Movimientos", "msg": str(e)})
            continue
        try:
            ff = _parse_date(row.get("Fecha_Fin"))
        except ValueError as e:
            errors.append({"row": rn, "sheet": "Movimientos", "msg": str(e)})
            continue

        cant_env = int(_parse_num(row.get("Cantidad_Enviada"), 0))
        cant_rec = int(_parse_num(row.get("Cantidad_Recibida"), 0))
        tarifa = float(_parse_num(row.get("Tarifa"), 0))

        mov = {
            "id": str(uuid.uuid4()),
            "servicio_id": servicio_id,
            "persona_id": persona_id,
            "fecha_inicio": fi,
            "fecha_fin": ff,
            "cantidad_enviada": cant_env,
            "cantidad_recibida": cant_rec,
            "tarifa": tarifa,
            "observaciones": _clean_str(row.get("Observaciones")) or None,
        }
        movimientos_por_corte.setdefault(n_corte, []).append(mov)

    # ── validar materiales ──
    items_by_codigo = {r["codigo"].strip().lower(): {"id": str(r["id"]), "nombre": r["nombre"], "unidad": r["unidad_medida"]}
                       for r in await conn.fetch("SELECT id, codigo, nombre, unidad_medida FROM prod_inventario WHERE codigo IS NOT NULL AND codigo != ''")}
    materiales_por_corte = {}
    for row in materiales_data:
        rn = row["_row"]
        n_corte = _clean_str(row.get("N_Corte"))
        if not n_corte:
            errors.append({"row": rn, "sheet": "Materiales", "msg": "N_Corte es obligatorio"})
            continue
        if n_corte not in n_corte_map:
            errors.append({"row": rn, "sheet": "Materiales", "msg": f"N_Corte '{n_corte}' no existe en pestaña Registros"})
            continue

        codigo = _clean_str(row.get("Item_Codigo")).lower()
        if not codigo:
            errors.append({"row": rn, "sheet": "Materiales", "msg": "Item_Codigo es obligatorio"})
            continue
        item_info = items_by_codigo.get(codigo)
        if not item_info:
            errors.append({"row": rn, "sheet": "Materiales", "msg": f"Item con código '{row.get('Item_Codigo')}' no existe en inventario"})
            continue

        cantidad = float(_parse_num(row.get("Cantidad"), 0))
        if cantidad <= 0:
            errors.append({"row": rn, "sheet": "Materiales", "msg": "Cantidad debe ser mayor a 0"})
            continue

        materiales_por_corte.setdefault(n_corte, []).append({
            "item_id": item_info["id"],
            "cantidad": cantidad,
            "observaciones": _clean_str(row.get("Observaciones")) or None,
        })

    return {
        "errors": errors,
        "warnings": warnings,
        "registros": n_corte_map,
        "tallas": tallas_por_corte,
        "movimientos": movimientos_por_corte,
        "materiales": materiales_por_corte,
    }


# ─────────────── TEMPLATE DOWNLOAD ────────────────

@router.get("/registros/import-template")
async def download_template():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    example_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def style_header(ws, headers, example_row=None):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border
            ws.column_dimensions[cell.column_letter].width = max(15, len(h) + 4)
        if example_row:
            for c, v in enumerate(example_row, 1):
                cell = ws.cell(row=2, column=c, value=v)
                cell.fill = example_fill
                cell.border = thin_border

    # ── Pestaña Registros ──
    ws1 = wb.active
    ws1.title = "Registros"
    reg_headers = [
        "N_Corte", "Nombre_Modelo", "Marca", "Tipo", "Tela", "Entalle",
        "Hilo", "Hilo_Especifico", "Prendas_Total", "Estado_Actual",
        "Fecha_Inicio", "Fecha_Entrega", "Linea_Negocio", "Urgente", "Observaciones"
    ]
    reg_example = [
        "C-001", "Polo Básico", "Nike", "Polo", "Jersey", "Slim",
        "Color", "Negro", 500, "Para Costura",
        "01/04/2026", "30/04/2026", "Producción", "NO", "Lote de prueba"
    ]
    style_header(ws1, reg_headers, reg_example)

    # ── Pestaña Movimientos ──
    ws2 = wb.create_sheet("Movimientos")
    mov_headers = [
        "N_Corte", "Servicio", "Persona", "Fecha_Inicio", "Fecha_Fin",
        "Cantidad_Enviada", "Cantidad_Recibida", "Tarifa", "Observaciones"
    ]
    mov_example = [
        "C-001", "Costura", "María López", "05/04/2026", "15/04/2026",
        500, 495, 1.50, "Primer servicio"
    ]
    style_header(ws2, mov_headers, mov_example)

    # ── Pestaña Tallas ──
    ws3 = wb.create_sheet("Tallas")
    talla_headers = ["N_Corte"] + TALLA_COLS
    talla_example = ["C-001", "", 100, 150, 150, 100, "", "", "", "", "", "", "", "", ""]
    style_header(ws3, talla_headers, talla_example)

    # ── Pestaña Materiales ──
    ws4 = wb.create_sheet("Materiales")
    mat_headers = ["N_Corte", "Item_Codigo", "Item_Nombre", "Cantidad", "Unidad", "Observaciones"]
    mat_example = ["C-001", "AVI-018", "Cierre Azul", 338, "unidad", "Opcional"]
    style_header(ws4, mat_headers, mat_example)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_importacion_registros.xlsx"},
    )


# ─────────────── EXPORT REGISTRO AS TEMPLATE ─────────────────

@router.get("/registros/{registro_id}/export-template")
async def export_registro_template(registro_id: str):
    """Exporta un registro existente como Excel compatible con la plantilla de importación."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    pool = await get_pool()
    async with pool.acquire() as conn:
        reg = await conn.fetchrow("""
            SELECT r.*, m.nombre AS modelo_nombre,
                   ma.nombre AS marca_nombre, tp.nombre AS tipo_nombre,
                   en.nombre AS entalle_nombre, te.nombre AS tela_nombre,
                   he.nombre AS hilo_esp_nombre, hi.nombre AS hilo_nombre,
                   ln.nombre AS linea_nombre
            FROM prod_registros r
            LEFT JOIN prod_modelos m ON m.id = r.modelo_id
            LEFT JOIN prod_marcas ma ON ma.id = m.marca_id
            LEFT JOIN prod_tipos tp ON tp.id = m.tipo_id
            LEFT JOIN prod_entalles en ON en.id = m.entalle_id
            LEFT JOIN prod_telas te ON te.id = m.tela_id
            LEFT JOIN prod_hilos_especificos he ON he.id = COALESCE(m.hilo_especifico_id, r.hilo_especifico_id)
            LEFT JOIN prod_hilos hi ON hi.id = m.hilo_id
            LEFT JOIN finanzas2.cont_linea_negocio ln ON ln.id = r.linea_negocio_id
            WHERE r.id = $1
        """, registro_id)
        if not reg:
            raise HTTPException(404, "Registro no encontrado")

        # Parse modelo_manual for fallback values
        mm = None
        if reg.get('modelo_manual'):
            import json as _json
            raw = reg['modelo_manual']
            mm = _json.loads(raw) if isinstance(raw, str) else dict(raw)

        modelo_nombre = reg['modelo_nombre'] or (mm.get('nombre_modelo') if mm else None) or ''
        marca_nombre = reg['marca_nombre'] or (mm.get('marca_texto') if mm else None) or ''
        tipo_nombre = reg['tipo_nombre'] or (mm.get('tipo_texto') if mm else None) or ''
        tela_nombre = reg['tela_nombre'] or (mm.get('tela_texto') if mm else None) or ''
        entalle_nombre = reg['entalle_nombre'] or (mm.get('entalle_texto') if mm else None) or ''
        hilo_nombre = reg['hilo_nombre'] or (mm.get('hilo_texto') if mm else None) or ''
        hilo_esp_nombre = reg['hilo_esp_nombre'] or (mm.get('hilo_especifico_texto') if mm else None) or ''

        # Tallas
        tallas_rows = await conn.fetch(
            "SELECT t.nombre, rt.cantidad_real FROM prod_registro_tallas rt JOIN prod_tallas_catalogo t ON t.id = rt.talla_id WHERE rt.registro_id = $1",
            registro_id
        )
        talla_map = {r['nombre']: int(r['cantidad_real'] or 0) for r in tallas_rows}

        # Prendas total
        prendas = sum(talla_map.values()) or 0

        # Movimientos
        movs = await conn.fetch("""
            SELECT s.nombre AS servicio, p.nombre AS persona,
                   mp.fecha_inicio, mp.fecha_fin, mp.cantidad_enviada,
                   mp.cantidad_recibida, mp.tarifa_aplicada, mp.observaciones
            FROM prod_movimientos_produccion mp
            LEFT JOIN prod_servicios_produccion s ON s.id = mp.servicio_id
            LEFT JOIN prod_personas_produccion p ON p.id = mp.persona_id
            WHERE mp.registro_id = $1
            ORDER BY mp.created_at
        """, registro_id)

        # Materiales manuales
        materiales = await conn.fetch("""
            SELECT i.codigo, i.nombre, i.unidad_medida,
                   r.cantidad_requerida, r.observaciones
            FROM prod_registro_requerimiento_mp r
            JOIN prod_inventario i ON i.id = r.item_id
            WHERE r.registro_id = $1 AND r.origen = 'MANUAL'
            ORDER BY r.created_at
        """, registro_id)

    # ── Build Excel ──
    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    data_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def write_headers(ws, headers):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border
            ws.column_dimensions[cell.column_letter].width = max(15, len(h) + 4)

    def write_row(ws, row_num, values):
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=c, value=v)
            cell.fill = data_fill
            cell.border = thin_border

    def fmt_date(d):
        if d is None:
            return ''
        if hasattr(d, 'strftime'):
            return d.strftime('%d/%m/%Y')
        return str(d)

    # Pestaña Registros
    ws1 = wb.active
    ws1.title = "Registros"
    reg_headers = [
        "N_Corte", "Nombre_Modelo", "Marca", "Tipo", "Tela", "Entalle",
        "Hilo", "Hilo_Especifico", "Prendas_Total", "Estado_Actual",
        "Fecha_Inicio", "Fecha_Entrega", "Linea_Negocio", "Urgente", "Observaciones"
    ]
    write_headers(ws1, reg_headers)
    write_row(ws1, 2, [
        reg['n_corte'], modelo_nombre, marca_nombre, tipo_nombre,
        tela_nombre, entalle_nombre, hilo_nombre, hilo_esp_nombre,
        prendas, reg['estado'],
        fmt_date(reg.get('fecha_inicio_real')), fmt_date(reg.get('fecha_entrega_final')),
        reg['linea_nombre'] or '', 'SI' if reg['urgente'] else 'NO',
        reg.get('observaciones') or '',
    ])

    # Pestaña Movimientos
    ws2 = wb.create_sheet("Movimientos")
    mov_headers = [
        "N_Corte", "Servicio", "Persona", "Fecha_Inicio", "Fecha_Fin",
        "Cantidad_Enviada", "Cantidad_Recibida", "Tarifa", "Observaciones"
    ]
    write_headers(ws2, mov_headers)
    for i, mv in enumerate(movs, 2):
        write_row(ws2, i, [
            reg['n_corte'], mv['servicio'] or '', mv['persona'] or '',
            fmt_date(mv['fecha_inicio']), fmt_date(mv['fecha_fin']),
            int(mv['cantidad_enviada'] or 0), int(mv['cantidad_recibida'] or 0),
            float(mv['tarifa_aplicada'] or 0), mv['observaciones'] or '',
        ])

    # Pestaña Tallas
    ws3 = wb.create_sheet("Tallas")
    talla_headers = ["N_Corte"] + TALLA_COLS
    write_headers(ws3, talla_headers)
    write_row(ws3, 2, [reg['n_corte']] + [talla_map.get(t, '') for t in TALLA_COLS])

    # Pestaña Materiales
    ws4 = wb.create_sheet("Materiales")
    mat_headers = ["N_Corte", "Item_Codigo", "Item_Nombre", "Cantidad", "Unidad", "Observaciones"]
    write_headers(ws4, mat_headers)
    for i, mat in enumerate(materiales, 2):
        write_row(ws4, i, [
            reg['n_corte'], mat['codigo'] or '', mat['nombre'] or '',
            float(mat['cantidad_requerida'] or 0), mat['unidad_medida'] or '',
            mat['observaciones'] or '',
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"registro_{reg['n_corte']}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ─────────────── VALIDATE ─────────────────

@router.post("/registros/import-validate")
async def validate_import(file: UploadFile = File(...), user=Depends(get_current_user)):
    content = await file.read()
    try:
        registros_data, movimientos_data, tallas_data, materiales_data = _read_excel(content)
    except Exception as e:
        raise HTTPException(400, f"Error al leer Excel: {e}")

    pool = await get_pool()
    async with pool.acquire() as conn:
        result = await _validate(registros_data, movimientos_data, tallas_data, materiales_data, conn)

    regs = result["registros"]
    movs = result["movimientos"]
    talls = result["tallas"]
    mats = result["materiales"]

    # Build preview
    preview = []
    for nc, reg in regs.items():
        preview.append({
            "n_corte": nc,
            "nombre_modelo": reg["modelo_manual"].get("nombre_modelo") or "",
            "marca": reg["modelo_manual"].get("marca_texto") or "",
            "tipo": reg["modelo_manual"].get("tipo_texto") or "",
            "estado": reg["estado"],
            "prendas": reg["prendas_total"],
            "movimientos": len(movs.get(nc, [])),
            "tallas": len(talls.get(nc, [])),
            "materiales": len(mats.get(nc, [])),
        })

    return {
        "valid": len(result["errors"]) == 0,
        "total_registros": len(regs),
        "total_movimientos": sum(len(v) for v in movs.values()),
        "total_tallas": sum(len(v) for v in talls.values()),
        "total_materiales": sum(len(v) for v in mats.values()),
        "errors": result["errors"],
        "warnings": result["warnings"],
        "preview": preview,
    }


# ─────────────── EXECUTE IMPORT ─────────────────

@router.post("/registros/import-execute")
async def execute_import(file: UploadFile = File(...), empresa_id: int = Query(8), user=Depends(get_current_user)):
    content = await file.read()
    try:
        registros_data, movimientos_data, tallas_data, materiales_data = _read_excel(content)
    except Exception as e:
        raise HTTPException(400, f"Error al leer Excel: {e}")

    pool = await get_pool()
    async with pool.acquire() as conn:
        result = await _validate(registros_data, movimientos_data, tallas_data, materiales_data, conn)

        if result["errors"]:
            raise HTTPException(400, detail={"message": "Hay errores de validación", "errors": result["errors"]})

        regs = result["registros"]
        movs = result["movimientos"]
        talls = result["tallas"]
        mats = result["materiales"]

        async with conn.transaction():
            registros_creados = 0
            movimientos_creados = 0
            tallas_creadas = 0
            materiales_creados = 0

            for n_corte, reg in regs.items():
                # Build tallas JSON
                tallas_list = talls.get(n_corte, [])
                tallas_json = json.dumps([{"talla_id": t["talla_id"], "talla_nombre": t["talla_nombre"], "cantidad": t["cantidad"]} for t in tallas_list]) if tallas_list else "[]"

                modelo_manual_json = json.dumps(reg["modelo_manual"])
                fecha_creacion = datetime.now()
                fecha_inicio = reg["fecha_inicio"]
                fecha_entrega = reg["fecha_entrega"]

                # INSERT registro
                await conn.execute("""
                    INSERT INTO prod_registros (
                        id, n_corte, modelo_id, curva, estado, urgente,
                        hilo_especifico_id, tallas, distribucion_colores,
                        fecha_creacion, pt_item_id, empresa_id, observaciones,
                        linea_negocio_id, fecha_entrega_final, fecha_inicio_real,
                        modelo_manual
                    ) VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14,$15,$16,$17)
                """,
                    reg["id"], n_corte, None, "", reg["estado"], reg["urgente"],
                    None, tallas_json, "[]",
                    fecha_creacion, None, empresa_id, reg["observaciones"],
                    reg["linea_negocio_id"], fecha_entrega, fecha_inicio,
                    modelo_manual_json,
                )

                # INSERT tallas into prod_registro_tallas
                for t in tallas_list:
                    await conn.execute("""
                        INSERT INTO prod_registro_tallas (id, registro_id, talla_id, cantidad_programada, cantidad_real)
                        VALUES ($1, $2, $3, $4, $5)
                    """, str(uuid.uuid4()), reg["id"], t["talla_id"], t["cantidad"], t["cantidad"])
                    tallas_creadas += 1

                registros_creados += 1

                # INSERT movimientos
                for mov in movs.get(n_corte, []):
                    diferencia = (mov["cantidad_enviada"] or 0) - (mov["cantidad_recibida"] or 0)
                    costo = round((mov["cantidad_recibida"] or 0) * mov["tarifa"], 2)
                    avance = 100 if mov["fecha_fin"] else 0

                    await conn.execute("""
                        INSERT INTO prod_movimientos_produccion (
                            id, registro_id, servicio_id, persona_id,
                            cantidad_enviada, cantidad_recibida, diferencia,
                            costo_calculado, tarifa_aplicada,
                            fecha_inicio, fecha_fin, fecha_esperada_movimiento,
                            responsable_movimiento, observaciones,
                            avance_porcentaje, avance_updated_at, created_at, detalle_costos
                        ) VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14,$15,$16,$17,$18)
                    """,
                        mov["id"], reg["id"], mov["servicio_id"], mov["persona_id"],
                        mov["cantidad_enviada"], mov["cantidad_recibida"], diferencia,
                        costo, mov["tarifa"],
                        mov["fecha_inicio"], mov["fecha_fin"], None,
                        None, mov["observaciones"],
                        avance, datetime.now() if avance else None,
                        datetime.now(), None,
                    )
                    movimientos_creados += 1

                # INSERT materiales manuales
                for mat in mats.get(n_corte, []):
                    await conn.execute("""
                        INSERT INTO prod_registro_requerimiento_mp (
                            id, registro_id, item_id, talla_id,
                            cantidad_requerida, cantidad_reservada, cantidad_consumida,
                            estado, empresa_id, origen, observaciones
                        ) VALUES ($1,$2,$3,NULL,$4,0,0,'PENDIENTE',$5,'MANUAL',$6)
                    """,
                        str(uuid.uuid4()), reg["id"], mat["item_id"],
                        mat["cantidad"], empresa_id, mat.get("observaciones", ""),
                    )
                    materiales_creados += 1

    return {
        "success": True,
        "registros_creados": registros_creados,
        "movimientos_creados": movimientos_creados,
        "tallas_creadas": tallas_creadas,
        "materiales_creados": materiales_creados,
    }
