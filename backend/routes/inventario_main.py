"""Router for inventory management endpoints (items, ingresos, salidas, ajustes, rollos, reservas, reconciliar)."""
import json
import uuid
from datetime import datetime, timezone
from fastapi import APIRouter, HTTPException, Depends, Query
from db import get_pool
from auth_utils import get_current_user
from models import (
    ItemInventarioCreate, IngresoInventarioCreate,
    SalidaInventarioCreate, AjusteInventarioCreate,
    IngresoInventario, SalidaInventario, AjusteInventario, ItemInventario,
)
from helpers import registrar_actividad, row_to_dict, parse_jsonb, validar_registro_activo
from routes.auditoria import audit_log_safe, get_usuario
from typing import Optional, List
from pydantic import BaseModel

CATEGORIAS_INVENTARIO = ["Telas", "Avios", "Otros"]

router = APIRouter(prefix="/api")

PREFIJOS_CATEGORIA = {
    "Avios": "AVI",
    "Telas": "TEL",
    "PT": "PT",
    "Servicios": "SRV",
    "Otros": "OTR",
}

@router.get("/inventario-categorias")
async def get_categorias():
    return {"categorias": CATEGORIAS_INVENTARIO}

@router.get("/inventario/siguiente-codigo")
async def get_siguiente_codigo(categoria: str = "Otros"):
    prefijo = PREFIJOS_CATEGORIA.get(categoria, "OTR")
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            "SELECT codigo FROM prod_inventario WHERE codigo LIKE $1 ORDER BY codigo DESC",
            f"{prefijo}-%"
        )
        max_num = 0
        for r in rows:
            try:
                num = int(r["codigo"].split("-")[1])
                if num > max_num:
                    max_num = num
            except (IndexError, ValueError):
                continue
        return {"codigo": f"{prefijo}-{str(max_num + 1).zfill(3)}"}

@router.get("/inventario")
async def get_inventario(
    limit: int = 50,
    offset: int = 0,
    search: str = "",
    categoria: str = "",
    stock_status: str = "",
    all: str = "",
    linea_negocio_id: str = "",
):
    pool = await get_pool()
    async with pool.acquire() as conn:
        conditions = []
        params = []
        param_idx = 1

        if search:
            conditions.append(f"(i.nombre ILIKE ${param_idx} OR i.codigo ILIKE ${param_idx})")
            params.append(f"%{search}%")
            param_idx += 1

        if categoria:
            conditions.append(f"i.categoria = ${param_idx}")
            params.append(categoria)
            param_idx += 1

        if linea_negocio_id:
            if linea_negocio_id == "global":
                conditions.append("i.linea_negocio_id IS NULL")
            else:
                conditions.append(f"(i.linea_negocio_id = ${param_idx} OR i.linea_negocio_id IS NULL)")
                params.append(int(linea_negocio_id))
                param_idx += 1

        where_clause = " AND ".join(conditions) if conditions else "TRUE"

        base_query = f"""
            FROM prod_inventario i
            LEFT JOIN LATERAL (
                SELECT SUM(rl.cantidad_reservada - rl.cantidad_liberada) as total_reservado
                FROM prod_inventario_reservas_linea rl
                JOIN prod_inventario_reservas r ON rl.reserva_id = r.id
                WHERE rl.item_id = i.id AND r.estado = 'ACTIVA'
            ) res ON true
            LEFT JOIN LATERAL (
                SELECT SUM(cantidad_disponible * costo_unitario) as valorizado
                FROM prod_inventario_ingresos
                WHERE item_id = i.id AND cantidad_disponible > 0
            ) val ON true
            WHERE {where_clause}
        """

        # stock_status filter is applied post-query since it depends on computed fields
        # But we can push it into SQL for efficiency
        if stock_status == 'sin_stock':
            base_query += " AND i.stock_actual <= 0"
        elif stock_status == 'stock_bajo':
            base_query += " AND i.stock_actual > 0 AND i.stock_actual <= i.stock_minimo"
        elif stock_status == 'ok':
            base_query += " AND i.stock_actual > i.stock_minimo"

        # Count
        count_row = await conn.fetchrow(f"SELECT COUNT(*) as total {base_query}", *params)
        total = count_row['total']

        select_fields = """
            SELECT i.*,
                COALESCE(res.total_reservado, 0) as total_reservado,
                COALESCE(val.valorizado, 0) as valorizado
        """

        order_clause = " ORDER BY i.nombre ASC"

        if all == "true":
            rows = await conn.fetch(f"{select_fields} {base_query} {order_clause}", *params)
        else:
            rows = await conn.fetch(
                f"{select_fields} {base_query} {order_clause} LIMIT ${param_idx} OFFSET ${param_idx + 1}",
                *params, limit, offset
            )

        result = []
        for r in rows:
            d = row_to_dict(r)
            d['total_reservado'] = float(d.get('total_reservado') or 0)
            d['stock_disponible'] = max(0, float(d.get('stock_actual', 0)) - d['total_reservado'])
            d['valorizado'] = float(d.get('valorizado') or 0)
            result.append(d)

        if all == "true":
            return result
        return {"items": result, "total": total, "limit": limit, "offset": offset}

@router.get("/inventario-filtros")
async def get_inventario_filtros():
    """Retorna categorias y unidades únicas para filtros."""
    pool = await get_pool()
    async with pool.acquire() as conn:
        cats = await conn.fetch("SELECT DISTINCT categoria FROM prod_inventario WHERE categoria IS NOT NULL ORDER BY categoria")
        return {
            "categorias": [r['categoria'] for r in cats],
        }


@router.get("/inventario/alertas-stock")
async def get_alertas_stock(
    modo: str = "fisico",
    incluir_ignorados: str = "false",
):
    """Retorna items con stock bajo o sin stock.
    modo: 'fisico' (stock_actual) o 'disponible' (stock_actual - reservado)
    Solo items con stock_minimo > 0 configurado.
    """
    pool = await get_pool()
    async with pool.acquire() as conn:
        ignorar_filter = "" if incluir_ignorados == "true" else "AND COALESCE(i.ignorar_alerta_stock, false) = false"

        modo_mig = await conn.fetchval(
            "SELECT valor FROM prod_configuracion WHERE clave = 'modo_migracion'"
        )
        en_migracion = modo_mig == 'true'

        rows = await conn.fetch(f"""
            SELECT i.id, i.codigo, i.nombre, i.categoria, i.unidad_medida,
                   i.stock_actual, i.stock_minimo, i.tipo_item,
                   COALESCE(i.ignorar_alerta_stock, false) as ignorar_alerta_stock,
                   COALESCE((
                       SELECT SUM(rl.cantidad_reservada - rl.cantidad_liberada)
                       FROM prod_inventario_reservas_linea rl
                       JOIN prod_inventario_reservas r ON rl.reserva_id = r.id
                       WHERE rl.item_id = i.id AND r.estado = 'ACTIVA'
                   ), 0) as total_reservado
            FROM prod_inventario i
            WHERE i.stock_minimo > 0
              {ignorar_filter}
            ORDER BY i.stock_actual ASC, i.nombre ASC
        """)
        
        result = []
        for r in rows:
            d = row_to_dict(r)
            stock_actual = float(d.get('stock_actual') or 0)
            total_reservado = float(d.get('total_reservado') or 0)
            stock_disponible = max(0, stock_actual - total_reservado)
            stock_minimo = int(d.get('stock_minimo') or 0)
            
            # Determinar el stock de referencia según modo
            stock_ref = stock_disponible if modo == "disponible" else stock_actual
            
            if stock_ref <= stock_minimo:
                # Durante migración, stock negativo es temporal — no alertar
                if en_migracion and stock_actual < 0:
                    continue
                d['stock_actual'] = stock_actual
                d['total_reservado'] = total_reservado
                d['stock_disponible'] = stock_disponible
                d['faltante'] = max(0, stock_minimo - stock_ref)
                d['estado_stock'] = 'SIN_STOCK' if stock_actual <= 0 else 'STOCK_BAJO'
                result.append(d)
        
        sin_stock = sum(1 for i in result if i['estado_stock'] == 'SIN_STOCK')
        stock_bajo = sum(1 for i in result if i['estado_stock'] == 'STOCK_BAJO')
        
        return {
            "items": result,
            "total": len(result),
            "sin_stock": sin_stock,
            "stock_bajo": stock_bajo,
            "modo": modo,
        }



@router.get("/inventario/stock-por-linea")
async def get_stock_por_linea():
    """Retorna stock agrupado por item y línea de negocio."""
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch("""
            SELECT 
                i.id as item_id, i.codigo, i.nombre, i.unidad_medida, i.categoria,
                i.linea_negocio_id as item_linea_id,
                i.stock_actual,
                ln.nombre as linea_nombre,
                ln.codigo as linea_codigo,
                COALESCE(SUM(CASE WHEN ing.cantidad_disponible > 0 THEN ing.cantidad_disponible ELSE 0 END), 0) as stock_disponible_ingresos,
                COALESCE(SUM(CASE WHEN ing.cantidad_disponible > 0 THEN ing.cantidad_disponible * ing.costo_unitario ELSE 0 END), 0) as valorizado
            FROM prod_inventario i
            LEFT JOIN prod_inventario_ingresos ing ON ing.item_id = i.id AND ing.cantidad_disponible > 0
            LEFT JOIN finanzas2.cont_linea_negocio ln ON i.linea_negocio_id = ln.id
            WHERE i.stock_actual > 0 OR ing.cantidad_disponible > 0
            GROUP BY i.id, i.codigo, i.nombre, i.unidad_medida, i.categoria, i.linea_negocio_id, i.stock_actual, ln.nombre, ln.codigo
            ORDER BY COALESCE(ln.nombre, 'ZZZZZ'), i.nombre
        """)
        return [dict(r) for r in rows]


@router.put("/inventario/{item_id}/ignorar-alerta")
async def toggle_ignorar_alerta(item_id: str, _u=Depends(get_current_user)):
    """Activa/desactiva ignorar alertas de stock para un item."""
    pool = await get_pool()
    async with pool.acquire() as conn:
        item = await conn.fetchrow("SELECT id, ignorar_alerta_stock FROM prod_inventario WHERE id = $1", item_id)
        if not item:
            raise HTTPException(status_code=404, detail="Item no encontrado")
        
        nuevo_valor = not (item['ignorar_alerta_stock'] or False)
        await conn.execute(
            "UPDATE prod_inventario SET ignorar_alerta_stock = $1 WHERE id = $2",
            nuevo_valor, item_id
        )
        return {"id": item_id, "ignorar_alerta_stock": nuevo_valor}


# ==================== KARDEX GENERAL ====================

@router.get("/inventario/kardex-general")
async def get_kardex_general(
    fecha_inicio: str = Query(...),
    fecha_fin: str = Query(...),
    linea_negocio_id: Optional[str] = Query(None),
    categoria: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    formato: Optional[str] = Query(None),  # "xlsx" para exportar
    current_user: dict = Depends(get_current_user),
):
    """Kardex general consolidado por rango de fechas. Una fila por item."""
    from datetime import date as date_type
    import io

    from datetime import timedelta
    try:
        fi = date_type.fromisoformat(fecha_inicio)
        ff = date_type.fromisoformat(fecha_fin)
    except ValueError:
        raise HTTPException(status_code=400, detail="Formato de fecha inválido (use YYYY-MM-DD)")

    # ff_excl: día siguiente para comparar con < en lugar de <= en timestamps
    ff_excl = ff + timedelta(days=1)

    pool = await get_pool()
    async with pool.acquire() as conn:
        # Build WHERE conditions for prod_inventario
        conditions = ["i.activo = true", "i.tipo_item != 'SERVICIO'"]
        params: list = [fi, ff_excl]
        p = 3  # next param index

        if linea_negocio_id and linea_negocio_id != "todos":
            conditions.append(f"i.linea_negocio_id = ${p}")
            params.append(int(linea_negocio_id))
            p += 1

        if categoria and categoria != "todos":
            conditions.append(f"i.categoria = ${p}")
            params.append(categoria)
            p += 1

        if search:
            conditions.append(f"(LOWER(i.nombre) LIKE ${p} OR LOWER(i.codigo) LIKE ${p})")
            params.append(f"%{search.lower()}%")
            p += 1

        where_clause = " AND ".join(conditions)

        sql = f"""
        WITH
        ing_antes AS (
            SELECT item_id, COALESCE(SUM(cantidad), 0) AS total
            FROM prod_inventario_ingresos WHERE fecha < $1 GROUP BY item_id
        ),
        sal_antes AS (
            SELECT item_id, COALESCE(SUM(cantidad), 0) AS total
            FROM prod_inventario_salidas WHERE fecha < $1 GROUP BY item_id
        ),
        adj_antes AS (
            SELECT item_id,
                   COALESCE(SUM(CASE WHEN tipo='entrada' THEN cantidad ELSE -cantidad END), 0) AS total
            FROM prod_inventario_ajustes WHERE fecha < $1 GROUP BY item_id
        ),
        ing_rango AS (
            SELECT item_id, COALESCE(SUM(cantidad), 0) AS total
            FROM prod_inventario_ingresos
            WHERE fecha >= $1 AND fecha < $2 GROUP BY item_id
        ),
        sal_rango AS (
            SELECT item_id, COALESCE(SUM(cantidad), 0) AS total
            FROM prod_inventario_salidas
            WHERE fecha >= $1 AND fecha < $2 GROUP BY item_id
        ),
        adj_rango AS (
            SELECT item_id,
                   COALESCE(SUM(CASE WHEN tipo='entrada' THEN cantidad ELSE 0 END), 0) AS entradas,
                   COALESCE(SUM(CASE WHEN tipo='salida' THEN cantidad ELSE 0 END), 0) AS salidas
            FROM prod_inventario_ajustes
            WHERE fecha >= $1 AND fecha < $2 GROUP BY item_id
        )
        SELECT
            i.id, i.codigo, i.nombre, i.categoria,
            i.unidad_medida,
            COALESCE(i.costo_promedio, 0) AS costo_promedio,
            COALESCE(ln.nombre, '') AS linea_negocio_nombre,
            -- Saldo inicial
            COALESCE(ia.total, 0) - COALESCE(sa.total, 0) + COALESCE(aa.total, 0) AS saldo_inicial,
            -- Ingresos en rango (ingresos + ajustes entrada)
            COALESCE(ir.total, 0) + COALESCE(ar.entradas, 0) AS ingresos,
            -- Salidas en rango (salidas + ajustes salida)
            COALESCE(sr.total, 0) + COALESCE(ar.salidas, 0) AS salidas
        FROM prod_inventario i
        LEFT JOIN finanzas2.cont_linea_negocio ln ON ln.id = i.linea_negocio_id
        LEFT JOIN ing_antes ia ON ia.item_id = i.id
        LEFT JOIN sal_antes sa ON sa.item_id = i.id
        LEFT JOIN adj_antes aa ON aa.item_id = i.id
        LEFT JOIN ing_rango ir ON ir.item_id = i.id
        LEFT JOIN sal_rango sr ON sr.item_id = i.id
        LEFT JOIN adj_rango ar ON ar.item_id = i.id
        WHERE {where_clause}
          AND (
            COALESCE(ia.total, 0) - COALESCE(sa.total, 0) + COALESCE(aa.total, 0) > 0
            OR COALESCE(ir.total, 0) + COALESCE(ar.entradas, 0) > 0
            OR COALESCE(sr.total, 0) + COALESCE(ar.salidas, 0) > 0
          )
        ORDER BY i.categoria, i.nombre
        """

        rows = await conn.fetch(sql, *params)

        items = []
        for r in rows:
            saldo_inicial = float(r['saldo_inicial'] or 0)
            ingresos = float(r['ingresos'] or 0)
            salidas = float(r['salidas'] or 0)
            saldo_final = saldo_inicial + ingresos - salidas
            costo_prom = float(r['costo_promedio'] or 0)
            valor_saldo = round(saldo_final * costo_prom, 2)
            items.append({
                "id": r['id'],
                "codigo": r['codigo'],
                "nombre": r['nombre'],
                "categoria": r['categoria'] or '',
                "linea_negocio_nombre": r['linea_negocio_nombre'],
                "unidad_medida": r['unidad_medida'] or '',
                "saldo_inicial": round(saldo_inicial, 4),
                "ingresos": round(ingresos, 4),
                "salidas": round(salidas, 4),
                "saldo_final": round(saldo_final, 4),
                "costo_promedio": round(costo_prom, 4),
                "valor_saldo_final": valor_saldo,
            })

        if formato == "xlsx":
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from fastapi.responses import StreamingResponse

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Kardex General"

            # Title
            ws.merge_cells("A1:K1")
            ws["A1"] = f"Kardex General — {fecha_inicio} al {fecha_fin}"
            ws["A1"].font = Font(bold=True, size=13)
            ws["A1"].alignment = Alignment(horizontal="center")

            # Header row
            headers = ["Código", "Nombre", "Categoría", "Línea de Negocio", "Unidad",
                       "Saldo Inicial", "Ingresos", "Salidas", "Saldo Final", "Costo Prom.", "Valor Saldo Final"]
            header_fill = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
            thin = Side(style="thin")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col_idx, value=h)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = border

            # Data rows
            for row_idx, item in enumerate(items, 3):
                vals = [
                    item["codigo"], item["nombre"], item["categoria"],
                    item["linea_negocio_nombre"], item["unidad_medida"],
                    item["saldo_inicial"], item["ingresos"], item["salidas"],
                    item["saldo_final"], item["costo_promedio"], item["valor_saldo_final"],
                ]
                fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") if row_idx % 2 == 0 else None
                for col_idx, val in enumerate(vals, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.border = border
                    if fill:
                        cell.fill = fill
                    if col_idx >= 6:
                        cell.number_format = '#,##0.00##'

            # Totals row
            total_row = len(items) + 3
            ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
            ws.merge_cells(f"A{total_row}:E{total_row}")
            total_valor = sum(i["valor_saldo_final"] for i in items)
            for col_idx in range(1, 12):
                cell = ws.cell(row=total_row, column=col_idx)
                cell.border = border
                cell.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
            ws.cell(row=total_row, column=11, value=round(total_valor, 2)).font = Font(bold=True)
            ws.cell(row=total_row, column=11).number_format = '#,##0.00'

            # Column widths
            widths = [12, 32, 14, 20, 10, 14, 14, 14, 14, 13, 18]
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            filename = f"kardex_general_{fecha_inicio}_{fecha_fin}.xlsx"
            return StreamingResponse(
                buf,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )

        total_valor = sum(i["valor_saldo_final"] for i in items)
        return {
            "items": items,
            "total_items": len(items),
            "total_valor_saldo": round(total_valor, 2),
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin,
        }


@router.get("/inventario/{item_id}")
async def get_item_inventario(item_id: str):
    pool = await get_pool()
    async with pool.acquire() as conn:
        item = await conn.fetchrow("SELECT * FROM prod_inventario WHERE id = $1", item_id)
        if not item:
            raise HTTPException(status_code=404, detail="Item no encontrado")
        d = row_to_dict(item)
        # Lotes disponibles
        ingresos = await conn.fetch(
            "SELECT * FROM prod_inventario_ingresos WHERE item_id = $1 AND cantidad_disponible > 0 ORDER BY fecha ASC", item_id
        )
        d['lotes'] = [row_to_dict(i) for i in ingresos]
        # Rollos si aplica
        if d.get('control_por_rollos'):
            rollos = await conn.fetch(
                "SELECT * FROM prod_inventario_rollos WHERE item_id = $1 AND activo = true AND metraje_disponible > 0", item_id
            )
            d['rollos'] = [row_to_dict(r) for r in rollos]
        return d


@router.get("/inventario/{item_id}/reservas-detalle")
async def get_reservas_detalle_item(item_id: str):
    """
    Obtiene el detalle de reservas activas para un item,
    agrupado por registro (para ver qué registros tienen reservas)
    """
    pool = await get_pool()
    async with pool.acquire() as conn:
        item = await conn.fetchrow("SELECT * FROM prod_inventario WHERE id = $1", item_id)
        if not item:
            raise HTTPException(status_code=404, detail="Item no encontrado")
        
        # Obtener todas las líneas de reserva activas para este item
        rows = await conn.fetch("""
            SELECT 
                rl.id,
                rl.item_id,
                rl.talla_id,
                rl.cantidad_reservada,
                rl.cantidad_liberada,
                (rl.cantidad_reservada - rl.cantidad_liberada) as cantidad_activa,
                res.id as reserva_id,
                res.registro_id,
                res.estado as reserva_estado,
                res.fecha as reserva_fecha,
                reg.n_corte,
                reg.estado as registro_estado,
                m.nombre as modelo_nombre,
                tc.nombre as talla_nombre
            FROM prod_inventario_reservas_linea rl
            JOIN prod_inventario_reservas res ON rl.reserva_id = res.id
            JOIN prod_registros reg ON res.registro_id = reg.id
            LEFT JOIN prod_modelos m ON reg.modelo_id = m.id
            LEFT JOIN prod_tallas_catalogo tc ON rl.talla_id = tc.id
            WHERE rl.item_id = $1 AND res.estado = 'ACTIVA' AND (rl.cantidad_reservada - rl.cantidad_liberada) > 0
            ORDER BY res.fecha DESC
        """, item_id)
        
        # Agrupar por registro
        registros_map = {}
        for r in rows:
            reg_id = r['registro_id']
            if reg_id not in registros_map:
                registros_map[reg_id] = {
                    'registro_id': reg_id,
                    'n_corte': r['n_corte'],
                    'registro_estado': r['registro_estado'],
                    'modelo_nombre': r['modelo_nombre'],
                    'total_reservado': 0,
                    'lineas': []
                }
            registros_map[reg_id]['total_reservado'] += float(r['cantidad_activa'])
            registros_map[reg_id]['lineas'].append({
                'id': r['id'],
                'talla_nombre': r['talla_nombre'],
                'cantidad_reservada': float(r['cantidad_reservada']),
                'cantidad_liberada': float(r['cantidad_liberada']),
                'cantidad_activa': float(r['cantidad_activa']),
                'reserva_fecha': r['reserva_fecha'].isoformat() if r['reserva_fecha'] else None
            })
        
        total_reservado = sum(reg['total_reservado'] for reg in registros_map.values())
        
        return {
            'item_id': item_id,
            'item_codigo': item['codigo'],
            'item_nombre': item['nombre'],
            'stock_actual': float(item['stock_actual']),
            'total_reservado': total_reservado,
            'stock_disponible': max(0, float(item['stock_actual']) - total_reservado),
            'registros': list(registros_map.values())
        }


@router.post("/inventario")
async def create_item_inventario(input: ItemInventarioCreate, _u=Depends(get_current_user)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        existing = await conn.fetchrow("SELECT id FROM prod_inventario WHERE codigo = $1", input.codigo)
        if existing:
            raise HTTPException(status_code=400, detail="El código ya existe")
        item = ItemInventario(**input.model_dump())
        item.stock_actual = 0  # Siempre empieza en 0; el stock_inicial genera un ingreso formal
        await conn.execute(
            """INSERT INTO prod_inventario (id, codigo, nombre, descripcion, categoria, unidad_medida, stock_minimo, stock_actual, control_por_rollos, linea_negocio_id, created_at)
               VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11)""",
            item.id, item.codigo, item.nombre, item.descripcion, item.categoria, item.unidad_medida,
            item.stock_minimo, 0, item.control_por_rollos, item.linea_negocio_id, item.created_at.replace(tzinfo=None)
        )
        # Si viene stock_inicial > 0, generar ingreso formal de tipo 'stock_inicial'
        if input.stock_inicial and input.stock_inicial > 0:
            ingreso_id = str(uuid.uuid4())
            costo_u = input.costo_unitario_inicial or 0
            await conn.execute(
                """INSERT INTO prod_inventario_ingresos
                       (id, item_id, cantidad, cantidad_disponible, costo_unitario, proveedor,
                        numero_documento, observaciones, tipo_ingreso, fecha, empresa_id, linea_negocio_id)
                   VALUES ($1,$2,$3,$4,$5,'','','Stock inicial cargado al crear el item','stock_inicial',$6,$7,$8)""",
                ingreso_id, item.id, input.stock_inicial, input.stock_inicial, costo_u,
                item.created_at.replace(tzinfo=None), 7, item.linea_negocio_id,
            )
            await conn.execute(
                "UPDATE prod_inventario SET stock_actual = $1 WHERE id = $2",
                input.stock_inicial, item.id,
            )
            if costo_u > 0:
                await conn.execute(
                    "UPDATE prod_inventario SET costo_promedio = $1 WHERE id = $2",
                    costo_u, item.id,
                )
            item.stock_actual = input.stock_inicial
        return item

@router.put("/inventario/{item_id}")
async def update_item_inventario(item_id: str, input: ItemInventarioCreate, _u=Depends(get_current_user)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        result = await conn.fetchrow("SELECT * FROM prod_inventario WHERE id = $1", item_id)
        if not result:
            raise HTTPException(status_code=404, detail="Item no encontrado")
        if input.codigo != result['codigo']:
            existing = await conn.fetchrow("SELECT id FROM prod_inventario WHERE codigo = $1 AND id != $2", input.codigo, item_id)
            if existing:
                raise HTTPException(status_code=400, detail="El código ya existe")
        await conn.execute(
            """UPDATE prod_inventario SET codigo=$1, nombre=$2, descripcion=$3, categoria=$4, unidad_medida=$5, stock_minimo=$6, control_por_rollos=$7, linea_negocio_id=$8 WHERE id=$9""",
            input.codigo, input.nombre, input.descripcion, input.categoria, input.unidad_medida, input.stock_minimo, input.control_por_rollos, input.linea_negocio_id, item_id
        )
        return {**row_to_dict(result), **input.model_dump()}

@router.delete("/inventario/{item_id}")
async def delete_item_inventario(item_id: str, _u=Depends(get_current_user)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        # Validar que no tenga movimientos
        tiene_salidas = await conn.fetchval("SELECT COUNT(*) FROM prod_inventario_salidas WHERE item_id = $1", item_id)
        if tiene_salidas > 0:
            raise HTTPException(status_code=400, detail="No se puede eliminar: el item tiene salidas registradas")
        tiene_ajustes = await conn.fetchval("SELECT COUNT(*) FROM prod_inventario_ajustes WHERE item_id = $1", item_id)
        if tiene_ajustes > 0:
            raise HTTPException(status_code=400, detail="No se puede eliminar: el item tiene ajustes registrados")
        tiene_reservas = await conn.fetchval(
            "SELECT COUNT(*) FROM prod_inventario_reservas_linea rl JOIN prod_inventario_reservas r ON r.id = rl.reserva_id WHERE rl.item_id = $1 AND r.estado = 'ACTIVA' AND rl.cantidad_reservada > rl.cantidad_liberada",
            item_id
        )
        if tiene_reservas > 0:
            raise HTTPException(status_code=400, detail="No se puede eliminar: el item tiene reservas activas")
        await conn.execute("DELETE FROM prod_inventario_ingresos WHERE item_id = $1", item_id)
        await conn.execute("DELETE FROM prod_inventario_rollos WHERE item_id = $1", item_id)
        await conn.execute("DELETE FROM prod_inventario WHERE id = $1", item_id)
        return {"message": "Item eliminado"}


# ==================== ENDPOINTS INGRESOS INVENTARIO ====================

@router.get("/inventario-ingresos")
async def get_ingresos():
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch("""
            SELECT ing.*,
                COALESCE(inv.nombre, '') as item_nombre,
                COALESCE(inv.codigo, '') as item_codigo,
                COALESCE(ln.nombre, '') as linea_negocio_nombre,
                COALESCE(rol.cnt, 0) as rollos_count,
                COALESCE(fac.qty_facturada, 0) as qty_facturada
            FROM prod_inventario_ingresos ing
            LEFT JOIN prod_inventario inv ON ing.item_id = inv.id
            LEFT JOIN finanzas2.cont_linea_negocio ln ON ln.id = ing.linea_negocio_id
            LEFT JOIN LATERAL (
                SELECT COUNT(*) as cnt FROM prod_inventario_rollos WHERE ingreso_id = ing.id
            ) rol ON true
            LEFT JOIN LATERAL (
                SELECT COALESCE(SUM(cantidad_aplicada), 0) as qty_facturada
                FROM finanzas2.cont_factura_ingreso_mp WHERE ingreso_id = ing.id
            ) fac ON true
            ORDER BY ing.fecha DESC
        """)
        result = []
        for r in rows:
            d = row_to_dict(r)
            qty_facturada = float(d.get('qty_facturada') or 0)
            qty_recibida = float(d.get('cantidad') or 0)
            d['qty_facturada'] = round(qty_facturada, 4)
            d['qty_pendiente_factura'] = round(max(0, qty_recibida - qty_facturada), 4)
            d['estado_facturacion'] = (
                'COMPLETO' if qty_recibida > 0 and (qty_recibida - qty_facturada) <= 0 else
                'PARCIAL' if qty_facturada > 0 else
                'PENDIENTE'
            )
            result.append(d)
        return result

@router.post("/inventario-ingresos")
async def create_ingreso(input: IngresoInventarioCreate, current_user: dict = Depends(get_current_user)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        item = await conn.fetchrow("SELECT * FROM prod_inventario WHERE id = $1", input.item_id)
        if not item:
            raise HTTPException(status_code=404, detail="Item de inventario no encontrado")
        
        rollos_data = input.rollos if hasattr(input, 'rollos') else []
        cantidad = input.cantidad
        
        if item['control_por_rollos'] and rollos_data:
            cantidad = sum(r.get('metraje', 0) for r in rollos_data)
        
        ingreso = IngresoInventario(
            item_id=input.item_id, cantidad=cantidad, costo_unitario=input.costo_unitario,
            proveedor=input.proveedor, numero_documento=input.numero_documento, observaciones=input.observaciones
        )
        ingreso.cantidad_disponible = cantidad
        
        # Línea de negocio: el item manda. Si el item tiene línea, se usa esa siempre.
        # Si el item es global (null), se usa lo que envíe el frontend (o null).
        linea_negocio_id = item.get('linea_negocio_id') or input.linea_negocio_id
        
        await conn.execute(
            """INSERT INTO prod_inventario_ingresos (id, item_id, cantidad, cantidad_disponible, costo_unitario, proveedor, numero_documento, observaciones, fecha, empresa_id, linea_negocio_id)
               VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11)""",
            ingreso.id, ingreso.item_id, ingreso.cantidad, ingreso.cantidad_disponible, ingreso.costo_unitario,
            ingreso.proveedor, ingreso.numero_documento, ingreso.observaciones, ingreso.fecha.replace(tzinfo=None), input.empresa_id, linea_negocio_id
        )
        
        # Crear rollos si aplica
        if item['control_por_rollos'] and rollos_data:
            for rollo_data in rollos_data:
                rollo_id = str(uuid.uuid4())
                await conn.execute(
                    """INSERT INTO prod_inventario_rollos (id, item_id, ingreso_id, numero_rollo, metraje, metraje_disponible, ancho, tono, observaciones, activo, created_at, empresa_id)
                       VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12)""",
                    rollo_id, input.item_id, ingreso.id, rollo_data.get('numero_rollo', ''), rollo_data.get('metraje', 0),
                    rollo_data.get('metraje', 0), rollo_data.get('ancho', 0), rollo_data.get('tono', ''),
                    rollo_data.get('observaciones', ''), True, datetime.now(timezone.utc).replace(tzinfo=None), input.empresa_id
                )
        
        # Actualizar stock
        await conn.execute("UPDATE prod_inventario SET stock_actual = stock_actual + $1 WHERE id = $2", cantidad, input.item_id)
        
        # Actualizar costo promedio ponderado desde ingresos disponibles
        await conn.execute("""
            UPDATE prod_inventario SET costo_promedio = COALESCE((
                SELECT SUM(cantidad_disponible * costo_unitario) / NULLIF(SUM(cantidad_disponible), 0)
                FROM prod_inventario_ingresos WHERE item_id = $1 AND cantidad_disponible > 0
            ), 0) WHERE id = $1
        """, input.item_id)
        
        await audit_log_safe(conn, get_usuario(current_user), "CREATE", "inventario", "prod_inventario_ingresos", ingreso.id,
            datos_despues={"item_id": input.item_id, "cantidad": cantidad, "costo_unitario": input.costo_unitario,
                           "proveedor": input.proveedor, "linea_negocio_id": input.linea_negocio_id},
            linea_negocio_id=input.linea_negocio_id)
        item_row = await conn.fetchrow("SELECT nombre FROM prod_inventario WHERE id = $1", input.item_id)
        item_nombre = item_row['nombre'] if item_row else input.item_id
    await registrar_actividad(pool, current_user['id'], current_user.get('username', ''), "crear",
        tabla_afectada="inventario", registro_id=ingreso.id, registro_nombre=item_nombre,
        descripcion=f"Ingreso de {cantidad} uds de {item_nombre}")
    return ingreso

class IngresoUpdateData(BaseModel):
    proveedor: str = ""
    numero_documento: str = ""
    observaciones: str = ""
    costo_unitario: float = 0
    rollos: Optional[List[dict]] = None


@router.get("/inventario-ingresos/ultimo-costo/{item_id}")
async def get_ultimo_costo_ingreso(item_id: str):
    """Retorna el costo unitario del último ingreso de un item."""
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT costo_unitario, fecha, proveedor FROM prod_inventario_ingresos WHERE item_id = $1 ORDER BY fecha DESC LIMIT 1",
            item_id
        )
        if not row:
            return {"tiene_historial": False, "costo_unitario": 0}
        return {
            "tiene_historial": True,
            "costo_unitario": float(row['costo_unitario'] or 0),
            "fecha": str(row['fecha']) if row['fecha'] else None,
            "proveedor": row['proveedor'],
        }


@router.get("/inventario-ingresos/{ingreso_id}/rollos")
async def get_ingreso_rollos(ingreso_id: str):
    pool = await get_pool()
    async with pool.acquire() as conn:
        ingreso = await conn.fetchrow("SELECT * FROM prod_inventario_ingresos WHERE id = $1", ingreso_id)
        if not ingreso:
            raise HTTPException(status_code=404, detail="Ingreso no encontrado")
        rollos = await conn.fetch(
            "SELECT id, numero_rollo, metraje, ancho, tono, estado FROM prod_inventario_rollos WHERE ingreso_id = $1 ORDER BY created_at ASC",
            ingreso_id
        )
        return [row_to_dict(r) for r in rollos]

@router.put("/inventario-ingresos/{ingreso_id}")
async def update_ingreso(ingreso_id: str, input: IngresoUpdateData, _u=Depends(get_current_user)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        ingreso = await conn.fetchrow("SELECT * FROM prod_inventario_ingresos WHERE id = $1", ingreso_id)
        if not ingreso:
            raise HTTPException(status_code=404, detail="Ingreso no encontrado")

        item = await conn.fetchrow("SELECT * FROM prod_inventario WHERE id = $1", ingreso['item_id'])

        # Actualizar campos base
        await conn.execute(
            """UPDATE prod_inventario_ingresos SET proveedor=$1, numero_documento=$2, observaciones=$3, costo_unitario=$4 WHERE id=$5""",
            input.proveedor, input.numero_documento, input.observaciones, input.costo_unitario, ingreso_id
        )

        # Si el item tiene control por rollos y se envían rollos, sincronizar
        if item and item['control_por_rollos'] and input.rollos is not None:
            old_cantidad = float(ingreso['cantidad'])

            # IDs de rollos que vienen del frontend (los que ya existían)
            incoming_ids = {r.get('id') for r in input.rollos if r.get('id')}

            # Eliminar rollos que ya no están (solo si no tienen movimientos)
            existing_rollos = await conn.fetch(
                "SELECT id, metraje, metraje_disponible FROM prod_inventario_rollos WHERE ingreso_id = $1", ingreso_id
            )
            for er in existing_rollos:
                if er['id'] not in incoming_ids:
                    tiene_salidas = await conn.fetchval(
                        "SELECT COUNT(*) FROM prod_inventario_salidas WHERE rollo_id = $1", er['id']
                    )
                    if tiene_salidas > 0:
                        raise HTTPException(status_code=400, detail=f"No se puede eliminar el rollo porque ya tiene salidas registradas")
                    if float(er['metraje']) != float(er['metraje_disponible']):
                        raise HTTPException(status_code=400, detail=f"No se puede eliminar el rollo porque ya tiene movimientos")
                    await conn.execute("DELETE FROM prod_inventario_rollos WHERE id = $1", er['id'])

            # Upsert rollos
            new_cantidad = 0
            for rollo_data in input.rollos:
                metraje = float(rollo_data.get('metraje', 0) or 0)
                ancho = float(rollo_data.get('ancho', 0) or 0)
                new_cantidad += metraje

                if rollo_data.get('id') and rollo_data['id'] in {er['id'] for er in existing_rollos}:
                    # Actualizar existente
                    await conn.execute(
                        """UPDATE prod_inventario_rollos 
                           SET numero_rollo=$1, metraje=$2, metraje_disponible=$2, ancho=$3, tono=$4
                           WHERE id=$5""",
                        rollo_data.get('numero_rollo', ''), metraje, ancho,
                        rollo_data.get('tono', ''), rollo_data['id']
                    )
                else:
                    # Crear nuevo
                    rollo_id = str(uuid.uuid4())
                    await conn.execute(
                        """INSERT INTO prod_inventario_rollos 
                           (id, item_id, ingreso_id, numero_rollo, metraje, metraje_disponible, ancho, tono, observaciones, activo, created_at, empresa_id)
                           VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12)""",
                        rollo_id, ingreso['item_id'], ingreso_id,
                        rollo_data.get('numero_rollo', ''), metraje, metraje,
                        ancho, rollo_data.get('tono', ''), '', True, datetime.now(timezone.utc).replace(tzinfo=None), ingreso['empresa_id']
                    )

            # Actualizar cantidad del ingreso y stock
            diff = new_cantidad - old_cantidad
            await conn.execute(
                "UPDATE prod_inventario_ingresos SET cantidad=$1, cantidad_disponible=cantidad_disponible+$2 WHERE id=$3",
                new_cantidad, diff, ingreso_id
            )
            await conn.execute(
                "UPDATE prod_inventario SET stock_actual = stock_actual + $1 WHERE id = $2",
                diff, ingreso['item_id']
            )

        return {"message": "Ingreso actualizado"}

@router.delete("/inventario-ingresos/{ingreso_id}")
async def delete_ingreso(ingreso_id: str, _u=Depends(get_current_user)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        ingreso = await conn.fetchrow("SELECT * FROM prod_inventario_ingresos WHERE id = $1", ingreso_id)
        if not ingreso:
            raise HTTPException(status_code=404, detail="Ingreso no encontrado")
        if ingreso['cantidad_disponible'] != ingreso['cantidad']:
            raise HTTPException(status_code=400, detail="No se puede eliminar un ingreso que ya tiene salidas")
        await conn.execute("DELETE FROM prod_inventario_rollos WHERE ingreso_id = $1", ingreso_id)
        await conn.execute("DELETE FROM prod_inventario_ingresos WHERE id = $1", ingreso_id)
        await conn.execute("UPDATE prod_inventario SET stock_actual = stock_actual - $1 WHERE id = $2", ingreso['cantidad'], ingreso['item_id'])
        return {"message": "Ingreso eliminado"}

# ==================== ENDPOINTS SALIDAS INVENTARIO ====================

@router.get("/inventario-salidas")
async def get_salidas(registro_id: str = None):
    pool = await get_pool()
    async with pool.acquire() as conn:
        where = "WHERE s.registro_id = $1" if registro_id else ""
        params = [registro_id] if registro_id else []
        rows = await conn.fetch(f"""
            SELECT s.*,
                COALESCE(inv.nombre, '') as item_nombre,
                COALESCE(inv.codigo, '') as item_codigo,
                COALESCE(ln.nombre, '') as linea_negocio_nombre,
                reg.n_corte as registro_n_corte,
                -- Nombre del modelo del registro (catálogo o manual) — útil para filtrar
                COALESCE(mod.nombre, reg.modelo_manual->>'nombre_modelo', '') as registro_modelo_nombre,
                rol.numero_rollo as rollo_numero,
                tal.nombre as talla_nombre
            FROM prod_inventario_salidas s
            LEFT JOIN prod_inventario inv ON s.item_id = inv.id
            LEFT JOIN finanzas2.cont_linea_negocio ln ON ln.id = s.linea_negocio_id
            LEFT JOIN prod_registros reg ON s.registro_id = reg.id
            LEFT JOIN prod_modelos mod ON mod.id = reg.modelo_id
            LEFT JOIN prod_inventario_rollos rol ON s.rollo_id = rol.id
            LEFT JOIN prod_tallas_catalogo tal ON s.talla_id = tal.id
            {where}
            ORDER BY s.fecha DESC
        """, *params)
        result = []
        for r in rows:
            d = row_to_dict(r)
            d['detalle_fifo'] = parse_jsonb(d.get('detalle_fifo'))
            result.append(d)
        return result

@router.post("/inventario-salidas")
async def create_salida(input: SalidaInventarioCreate, current_user: dict = Depends(get_current_user)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        item = await conn.fetchrow("SELECT * FROM prod_inventario WHERE id = $1", input.item_id)
        if not item:
            raise HTTPException(status_code=404, detail="Item de inventario no encontrado")
        
        control_por_rollos = item['control_por_rollos']
        
        # === FASE 2: Validaciones de reserva y rollo ===
        
        # Validar regla de rollo según control_por_rollos
        if control_por_rollos:
            if not input.rollo_id:
                # Auto-seleccionar rollo FIFO (el más antiguo con metraje disponible)
                auto_rollo = await conn.fetchrow("""
                    SELECT r.id, r.metraje_disponible FROM prod_inventario_rollos r
                    JOIN prod_inventario_ingresos ing ON r.ingreso_id = ing.id
                    WHERE r.item_id = $1 AND r.metraje_disponible > 0
                    ORDER BY ing.fecha ASC
                    LIMIT 1
                """, input.item_id)
                if not auto_rollo:
                    raise HTTPException(status_code=400, detail="No hay rollos disponibles para este item")
                if float(auto_rollo['metraje_disponible']) < input.cantidad:
                    raise HTTPException(status_code=400, detail=f"Metraje insuficiente en rollo disponible. Disponible: {auto_rollo['metraje_disponible']}")
                input.rollo_id = auto_rollo['id']
            # Validar que el rollo pertenece al item
            rollo = await conn.fetchrow("SELECT * FROM prod_inventario_rollos WHERE id = $1", input.rollo_id)
            if not rollo:
                raise HTTPException(status_code=404, detail="Rollo no encontrado")
            if rollo['item_id'] != input.item_id:
                raise HTTPException(status_code=400, detail="El rollo no pertenece a este item")
            if float(rollo['metraje_disponible']) < input.cantidad:
                raise HTTPException(status_code=400, detail=f"Metraje insuficiente en rollo. Disponible: {rollo['metraje_disponible']}")
        else:
            # NO TELA: rollo_id debe ser NULL
            if input.rollo_id:
                raise HTTPException(status_code=400, detail="Este item no usa control por rollos, rollo_id debe ser vacío")

        # Validar registro si se proporciona
        if input.registro_id:
            reg = await conn.fetchrow("SELECT * FROM prod_registros WHERE id = $1", input.registro_id)
            if not reg:
                raise HTTPException(status_code=404, detail="Registro no encontrado")

            validar_registro_activo(reg, contexto='crear salidas')

        # Validar stock suficiente (salvo modo migración global activo)
        if not control_por_rollos:
            modo_mig = await conn.fetchval(
                "SELECT valor FROM prod_configuracion WHERE clave = 'modo_migracion'"
            )
            if modo_mig != 'true' and float(item['stock_actual']) < input.cantidad:
                raise HTTPException(status_code=400, detail=f"Stock insuficiente. Disponible: {item['stock_actual']}")

            # Buscar requerimiento (informativo, no bloquea la salida)
            if input.talla_id:
                req = await conn.fetchrow("""
                    SELECT * FROM prod_registro_requerimiento_mp
                    WHERE registro_id = $1 AND item_id = $2 AND talla_id = $3
                """, input.registro_id, input.item_id, input.talla_id)
            else:
                req = await conn.fetchrow("""
                    SELECT * FROM prod_registro_requerimiento_mp
                    WHERE registro_id = $1 AND item_id = $2 AND talla_id IS NULL
                """, input.registro_id, input.item_id)
        
        # === FIN Validaciones Fase 2 ===
        
        costo_total = 0.0
        detalle_fifo = []

        if input.rollo_id:
            # Ya validamos el rollo arriba, lo obtenemos de nuevo para el ingreso
            rollo = await conn.fetchrow("SELECT * FROM prod_inventario_rollos WHERE id = $1", input.rollo_id)
            ingreso = await conn.fetchrow("SELECT costo_unitario FROM prod_inventario_ingresos WHERE id = $1", rollo['ingreso_id'])
            costo_unitario = float(ingreso['costo_unitario']) if ingreso else 0
            costo_total = input.cantidad * costo_unitario
            detalle_fifo = [{"rollo_id": input.rollo_id, "cantidad": input.cantidad, "costo_unitario": costo_unitario}]
            await conn.execute("UPDATE prod_inventario_rollos SET metraje_disponible = metraje_disponible - $1 WHERE id = $2", input.cantidad, input.rollo_id)
            await conn.execute("UPDATE prod_inventario_ingresos SET cantidad_disponible = cantidad_disponible - $1 WHERE id = $2", input.cantidad, rollo['ingreso_id'])
        else:
            ingresos = await conn.fetch(
                "SELECT * FROM prod_inventario_ingresos WHERE item_id = $1 AND cantidad_disponible > 0 ORDER BY fecha ASC", input.item_id
            )
            cantidad_restante = input.cantidad
            for ing in ingresos:
                if cantidad_restante <= 0:
                    break
                disponible = float(ing['cantidad_disponible'])
                consumir = min(disponible, cantidad_restante)
                costo_unitario = float(ing['costo_unitario'])
                costo_total += consumir * costo_unitario
                detalle_fifo.append({"ingreso_id": ing['id'], "cantidad": consumir, "costo_unitario": costo_unitario})
                await conn.execute("UPDATE prod_inventario_ingresos SET cantidad_disponible = cantidad_disponible - $1 WHERE id = $2", consumir, ing['id'])
                cantidad_restante -= consumir
        
        salida = SalidaInventario(**input.model_dump())
        salida.costo_total = costo_total
        salida.detalle_fifo = detalle_fifo
        
        # empresa_id: preferir del registro, luego del item, fallback 7
        empresa_id = 7
        # linea_negocio_id: heredar del registro si existe
        linea_negocio_id = input.linea_negocio_id
        if input.registro_id:
            reg = await conn.fetchrow("SELECT empresa_id, linea_negocio_id FROM prod_registros WHERE id = $1", input.registro_id)
            if reg and reg['empresa_id']:
                empresa_id = reg['empresa_id']
            if reg and reg['linea_negocio_id'] and not linea_negocio_id:
                linea_negocio_id = reg['linea_negocio_id']
        elif item.get('empresa_id'):
            empresa_id = item['empresa_id']
        
        # Insertar salida con talla_id, empresa_id y linea_negocio_id
        await conn.execute(
            """INSERT INTO prod_inventario_salidas (id, item_id, cantidad, registro_id, talla_id, observaciones, rollo_id, costo_total, detalle_fifo, fecha, empresa_id, linea_negocio_id)
               VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12)""",
            salida.id, salida.item_id, salida.cantidad, salida.registro_id, salida.talla_id, salida.observaciones,
            salida.rollo_id, salida.costo_total, json.dumps(salida.detalle_fifo), salida.fecha.replace(tzinfo=None),
            empresa_id, linea_negocio_id
        )
        await conn.execute("UPDATE prod_inventario SET stock_actual = stock_actual - $1 WHERE id = $2", input.cantidad, input.item_id)

        # === FASE 2: Actualizar cantidad_consumida en requerimiento ===
        if input.registro_id:
            if input.talla_id:
                await conn.execute("""
                    UPDATE prod_registro_requerimiento_mp
                    SET cantidad_consumida = cantidad_consumida + $1,
                        estado = CASE
                            WHEN cantidad_consumida + $1 >= cantidad_requerida THEN 'COMPLETO'
                            ELSE 'PARCIAL'
                        END,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE registro_id = $2 AND item_id = $3 AND talla_id = $4
                """, input.cantidad, input.registro_id, input.item_id, input.talla_id)
            else:
                await conn.execute("""
                    UPDATE prod_registro_requerimiento_mp
                    SET cantidad_consumida = cantidad_consumida + $1,
                        estado = CASE
                            WHEN cantidad_consumida + $1 >= cantidad_requerida THEN 'COMPLETO'
                            ELSE 'PARCIAL'
                        END,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE registro_id = $2 AND item_id = $3 AND talla_id IS NULL
                """, input.cantidad, input.registro_id, input.item_id)
            
            # Liberar TODA la reserva restante para este item/registro (la materia prima ya se consumió)
            reserva_row = await conn.fetchrow("""
                SELECT res.id as reserva_id FROM prod_inventario_reservas res
                WHERE res.registro_id = $1 AND res.estado = 'ACTIVA'
                ORDER BY res.fecha DESC LIMIT 1
            """, input.registro_id)
            if reserva_row:
                if input.talla_id:
                    await conn.execute("""
                        UPDATE prod_inventario_reservas_linea
                        SET cantidad_liberada = cantidad_reservada
                        WHERE reserva_id = $1 AND item_id = $2 AND talla_id = $3
                    """, reserva_row['reserva_id'], input.item_id, input.talla_id)
                else:
                    await conn.execute("""
                        UPDATE prod_inventario_reservas_linea
                        SET cantidad_liberada = cantidad_reservada
                        WHERE reserva_id = $1 AND item_id = $2 AND talla_id IS NULL
                    """, reserva_row['reserva_id'], input.item_id)
        
        await audit_log_safe(conn, get_usuario(current_user), "CREATE", "inventario", "prod_inventario_salidas", salida.id,
            datos_despues={"item_id": input.item_id, "cantidad": input.cantidad, "costo_total": round(costo_total, 4),
                           "registro_id": input.registro_id, "capas_fifo": len(detalle_fifo)},
            linea_negocio_id=linea_negocio_id, referencia=input.registro_id)
        item_row = await conn.fetchrow("SELECT nombre FROM prod_inventario WHERE id = $1", input.item_id)
        item_nombre = item_row['nombre'] if item_row else input.item_id
    await registrar_actividad(pool, current_user['id'], current_user.get('username', ''), "crear",
        tabla_afectada="inventario", registro_id=salida.id, registro_nombre=item_nombre,
        descripcion=f"Salida de {input.cantidad} uds de {item_nombre}")
    return salida


@router.post("/inventario/reconciliar-reservas")
async def reconciliar_reservas(_u=Depends(get_current_user)):
    """Sincroniza cantidad_liberada en reservas: si ya hubo salida para un item+registro, libera toda la reserva."""
    pool = await get_pool()
    async with pool.acquire() as conn:
        reservas_activas = await conn.fetch("""
            SELECT rl.id, rl.reserva_id, rl.item_id, rl.talla_id, rl.cantidad_reservada, rl.cantidad_liberada,
                   res.registro_id
            FROM produccion.prod_inventario_reservas_linea rl
            JOIN produccion.prod_inventario_reservas res ON rl.reserva_id = res.id
            WHERE res.estado = 'ACTIVA' AND (rl.cantidad_reservada - rl.cantidad_liberada) > 0
        """)
        corregidas = 0
        for rl in reservas_activas:
            if rl['talla_id']:
                total_salido = await conn.fetchval("""
                    SELECT COALESCE(SUM(cantidad), 0) FROM produccion.prod_inventario_salidas
                    WHERE item_id = $1 AND registro_id = $2 AND talla_id = $3
                """, rl['item_id'], rl['registro_id'], rl['talla_id'])
            else:
                total_salido = await conn.fetchval("""
                    SELECT COALESCE(SUM(cantidad), 0) FROM produccion.prod_inventario_salidas
                    WHERE item_id = $1 AND registro_id = $2 AND talla_id IS NULL
                """, rl['item_id'], rl['registro_id'])
            
            # Si ya hubo cualquier salida, liberar TODA la reserva
            if float(total_salido) > 0 and float(rl['cantidad_liberada']) < float(rl['cantidad_reservada']):
                await conn.execute("""
                    UPDATE produccion.prod_inventario_reservas_linea
                    SET cantidad_liberada = cantidad_reservada
                    WHERE id = $1
                """, rl['id'])
                corregidas += 1
        
        return {"message": f"Reconciliación completada. {corregidas} líneas corregidas."}



# OPCIÓN 2: Endpoint para Salida Extra (sin validación de reserva)
class SalidaExtraCreate(BaseModel):
    item_id: str
    cantidad: float
    registro_id: str
    talla_id: Optional[str] = None
    observaciones: str = ""
    rollo_id: Optional[str] = None
    motivo: str = "Consumo adicional"

@router.post("/inventario-salidas/extra")
async def create_salida_extra(input: SalidaExtraCreate, _u=Depends(get_current_user)):
    """
    Crea una salida SIN validar reserva previa.
    Útil para excedentes, reposiciones o ajustes.
    Solo valida stock/rollo disponible.
    """
    pool = await get_pool()
    async with pool.acquire() as conn:
        item = await conn.fetchrow("SELECT * FROM prod_inventario WHERE id = $1", input.item_id)
        if not item:
            raise HTTPException(status_code=404, detail="Item de inventario no encontrado")
        
        control_por_rollos = item['control_por_rollos']
        
        # Validar regla de rollo según control_por_rollos
        if control_por_rollos:
            if not input.rollo_id:
                raise HTTPException(status_code=400, detail="Este item requiere seleccionar un rollo")
            rollo = await conn.fetchrow("SELECT * FROM prod_inventario_rollos WHERE id = $1", input.rollo_id)
            if not rollo:
                raise HTTPException(status_code=404, detail="Rollo no encontrado")
            if rollo['item_id'] != input.item_id:
                raise HTTPException(status_code=400, detail="El rollo no pertenece a este item")
            if float(rollo['metraje_disponible']) < input.cantidad:
                raise HTTPException(status_code=400, detail=f"Metraje insuficiente en rollo. Disponible: {rollo['metraje_disponible']}")
        else:
            if input.rollo_id:
                raise HTTPException(status_code=400, detail="Este item no usa control por rollos")
            modo_mig = await conn.fetchval(
                "SELECT valor FROM prod_configuracion WHERE clave = 'modo_migracion'"
            )
            if modo_mig != 'true' and float(item['stock_actual']) < input.cantidad:
                raise HTTPException(status_code=400, detail=f"Stock insuficiente. Disponible: {item['stock_actual']}")

        # Validar registro
        if input.registro_id:
            reg = await conn.fetchrow("SELECT * FROM prod_registros WHERE id = $1", input.registro_id)
            if not reg:
                raise HTTPException(status_code=404, detail="Registro no encontrado")
            
            validar_registro_activo(reg, contexto='crear salidas')
        
        # NO validamos reserva - es salida extra
        
        costo_total = 0.0
        detalle_fifo = []
        
        if input.rollo_id:
            rollo = await conn.fetchrow("SELECT * FROM prod_inventario_rollos WHERE id = $1", input.rollo_id)
            ingreso = await conn.fetchrow("SELECT costo_unitario FROM prod_inventario_ingresos WHERE id = $1", rollo['ingreso_id'])
            costo_unitario = float(ingreso['costo_unitario']) if ingreso else 0
            costo_total = input.cantidad * costo_unitario
            detalle_fifo = [{"rollo_id": input.rollo_id, "cantidad": input.cantidad, "costo_unitario": costo_unitario}]
            await conn.execute("UPDATE prod_inventario_rollos SET metraje_disponible = metraje_disponible - $1 WHERE id = $2", input.cantidad, input.rollo_id)
            await conn.execute("UPDATE prod_inventario_ingresos SET cantidad_disponible = cantidad_disponible - $1 WHERE id = $2", input.cantidad, rollo['ingreso_id'])
        else:
            ingresos = await conn.fetch(
                "SELECT * FROM prod_inventario_ingresos WHERE item_id = $1 AND cantidad_disponible > 0 ORDER BY fecha ASC", input.item_id
            )
            cantidad_restante = input.cantidad
            for ing in ingresos:
                if cantidad_restante <= 0:
                    break
                disponible = float(ing['cantidad_disponible'])
                consumir = min(disponible, cantidad_restante)
                costo_unitario = float(ing['costo_unitario'])
                costo_total += consumir * costo_unitario
                detalle_fifo.append({"ingreso_id": ing['id'], "cantidad": consumir, "costo_unitario": costo_unitario})
                await conn.execute("UPDATE prod_inventario_ingresos SET cantidad_disponible = cantidad_disponible - $1 WHERE id = $2", consumir, ing['id'])
                cantidad_restante -= consumir
        
        salida_id = str(uuid.uuid4())
        fecha = datetime.now(timezone.utc)
        observaciones = f"[EXTRA] {input.motivo}. {input.observaciones}".strip()
        
        # empresa_id: preferir del registro, luego del item, fallback 7
        empresa_id = 7
        if input.registro_id:
            reg = await conn.fetchrow("SELECT empresa_id FROM prod_registros WHERE id = $1", input.registro_id)
            if reg and reg['empresa_id']:
                empresa_id = reg['empresa_id']
        elif item.get('empresa_id'):
            empresa_id = item['empresa_id']
        
        await conn.execute(
            """INSERT INTO prod_inventario_salidas (id, item_id, cantidad, registro_id, talla_id, observaciones, rollo_id, costo_total, detalle_fifo, fecha, empresa_id)
               VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11)""",
            salida_id, input.item_id, input.cantidad, input.registro_id, input.talla_id, observaciones,
            input.rollo_id, costo_total, json.dumps(detalle_fifo), fecha.replace(tzinfo=None),
            empresa_id
        )
        await conn.execute("UPDATE prod_inventario SET stock_actual = stock_actual - $1 WHERE id = $2", input.cantidad, input.item_id)
        
        # Actualizar requerimiento si existe (suma al consumido aunque no tenga reserva)
        if input.registro_id:
            if input.talla_id:
                await conn.execute("""
                    UPDATE prod_registro_requerimiento_mp
                    SET cantidad_consumida = cantidad_consumida + $1,
                        estado = CASE
                            WHEN cantidad_consumida + $1 >= cantidad_requerida THEN 'COMPLETO'
                            ELSE 'PARCIAL'
                        END,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE registro_id = $2 AND item_id = $3 AND talla_id = $4
                """, input.cantidad, input.registro_id, input.item_id, input.talla_id)
            else:
                await conn.execute("""
                    UPDATE prod_registro_requerimiento_mp
                    SET cantidad_consumida = cantidad_consumida + $1,
                        estado = CASE
                            WHEN cantidad_consumida + $1 >= cantidad_requerida THEN 'COMPLETO'
                            ELSE 'PARCIAL'
                        END,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE registro_id = $2 AND item_id = $3 AND talla_id IS NULL
                """, input.cantidad, input.registro_id, input.item_id)
        
        return {
            "id": salida_id,
            "item_id": input.item_id,
            "cantidad": input.cantidad,
            "costo_total": costo_total,
            "tipo": "EXTRA",
            "message": "Salida extra registrada"
        }


class SalidaUpdateData(BaseModel):
    observaciones: str = ""
    # Campos opcionales: sólo aplican en modo carga inicial (corregir ítem o cantidad).
    item_id: Optional[str] = None
    cantidad: Optional[float] = None

@router.put("/inventario-salidas/{salida_id}")
async def update_salida(salida_id: str, input: SalidaUpdateData, _u=Depends(get_current_user)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        salida = await conn.fetchrow("SELECT * FROM prod_inventario_salidas WHERE id = $1", salida_id)
        if not salida:
            raise HTTPException(status_code=404, detail="Salida no encontrada")

        # ¿Se está intentando cambiar el ítem o la cantidad?
        quiere_cambiar_item = input.item_id and input.item_id != salida['item_id']
        nueva_cantidad = float(input.cantidad) if input.cantidad is not None else None
        cantidad_vieja = float(salida['cantidad'])
        quiere_cambiar_cantidad = (
            nueva_cantidad is not None and abs(nueva_cantidad - cantidad_vieja) > 1e-6
        )

        if quiere_cambiar_item or quiere_cambiar_cantidad:
            # Solo permitido en modo carga inicial — en operación normal se debe
            # eliminar la salida y crear una nueva para que el FIFO sea correcto.
            modo_mig = await conn.fetchval(
                "SELECT valor FROM prod_configuracion WHERE clave = 'modo_migracion'"
            )
            if modo_mig != 'true':
                raise HTTPException(
                    status_code=400,
                    detail="Para cambiar el ítem o la cantidad de una salida, activa el Modo Carga Inicial. En operación normal elimina la salida y crea una nueva."
                )

            # No soportamos cambio con rollos (control FIFO por rollo es estricto).
            if salida.get('rollo_id'):
                raise HTTPException(
                    status_code=400,
                    detail="Esta salida está vinculada a un rollo. Debe eliminarse y recrearse para cambiar el ítem o la cantidad."
                )

            # 1) Revertir stock del ítem viejo (sumar la cantidad que se había restado)
            await conn.execute(
                "UPDATE prod_inventario SET stock_actual = stock_actual + $1 WHERE id = $2",
                cantidad_vieja, salida['item_id']
            )

            # 2) Determinar ítem y cantidad nuevos
            nuevo_item_id = input.item_id or salida['item_id']
            nuevo_cantidad = nueva_cantidad if nueva_cantidad is not None else cantidad_vieja

            # Validar que el ítem nuevo existe
            nuevo_item = await conn.fetchrow("SELECT * FROM prod_inventario WHERE id = $1", nuevo_item_id)
            if not nuevo_item:
                raise HTTPException(status_code=404, detail="Item nuevo no encontrado")

            # 3) Descontar stock del ítem nuevo
            await conn.execute(
                "UPDATE prod_inventario SET stock_actual = stock_actual - $1 WHERE id = $2",
                nuevo_cantidad, nuevo_item_id
            )

            # 4) Costo: usar costo promedio del nuevo ítem (simplificación en modo carga)
            costo_unitario = float(nuevo_item.get('costo_promedio') or 0)
            costo_total = round(costo_unitario * nuevo_cantidad, 4)

            # 5) Actualizar la salida
            await conn.execute(
                """UPDATE prod_inventario_salidas
                   SET item_id = $1, cantidad = $2, costo_unitario = $3, costo_total = $4,
                       detalle_fifo = NULL, observaciones = $5
                   WHERE id = $6""",
                nuevo_item_id, nuevo_cantidad, costo_unitario, costo_total,
                input.observaciones, salida_id
            )
            return {
                "message": "Salida actualizada (ítem/cantidad reemplazado)",
                "warning": "El detalle FIFO se limpió; el costo usa el promedio actual del nuevo ítem."
            }

        # Caso normal: sólo observaciones
        await conn.execute("UPDATE prod_inventario_salidas SET observaciones=$1 WHERE id=$2", input.observaciones, salida_id)
        return {"message": "Salida actualizada"}

@router.delete("/inventario-salidas/{salida_id}")
async def delete_salida(salida_id: str, _u=Depends(get_current_user)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        # Bloquear eliminación de salidas mientras el modo carga inicial esté activo
        modo_mig = await conn.fetchval(
            "SELECT valor FROM prod_configuracion WHERE clave = 'modo_migracion'"
        )
        if modo_mig == 'true':
            raise HTTPException(
                status_code=400,
                detail="No se puede eliminar salidas mientras el modo carga inicial esté activo. Desactívalo primero.",
            )
        salida = await conn.fetchrow("SELECT * FROM prod_inventario_salidas WHERE id = $1", salida_id)
        if not salida:
            raise HTTPException(status_code=404, detail="Salida no encontrada")
        # Bloquear eliminación de salidas ya revertidas por un período de migración cerrado
        if salida.get('revertida_por_migracion_id'):
            raise HTTPException(
                status_code=400,
                detail="Esta salida ya fue revertida por un período de carga inicial y no puede eliminarse.",
            )
        detalle_fifo = parse_jsonb(salida['detalle_fifo'])
        for detalle in detalle_fifo:
            if detalle.get('rollo_id'):
                await conn.execute("UPDATE prod_inventario_rollos SET metraje_disponible = metraje_disponible + $1 WHERE id = $2", detalle['cantidad'], detalle['rollo_id'])
                rollo = await conn.fetchrow("SELECT ingreso_id FROM prod_inventario_rollos WHERE id = $1", detalle['rollo_id'])
                if rollo:
                    await conn.execute("UPDATE prod_inventario_ingresos SET cantidad_disponible = cantidad_disponible + $1 WHERE id = $2", detalle['cantidad'], rollo['ingreso_id'])
            elif detalle.get('ingreso_id'):
                await conn.execute("UPDATE prod_inventario_ingresos SET cantidad_disponible = cantidad_disponible + $1 WHERE id = $2", detalle['cantidad'], detalle['ingreso_id'])
        await conn.execute("DELETE FROM prod_inventario_salidas WHERE id = $1", salida_id)
        await conn.execute("UPDATE prod_inventario SET stock_actual = stock_actual + $1 WHERE id = $2", float(salida['cantidad']), salida['item_id'])
        # Actualizar cantidad_consumida en requerimiento del registro
        if salida['registro_id']:
            talla_id = salida.get('talla_id')
            if talla_id:
                await conn.execute("""
                    UPDATE prod_registro_requerimiento_mp
                    SET cantidad_consumida = GREATEST(cantidad_consumida - $1, 0), updated_at = CURRENT_TIMESTAMP
                    WHERE registro_id = $2 AND item_id = $3 AND talla_id = $4
                """, float(salida['cantidad']), salida['registro_id'], salida['item_id'], talla_id)
            else:
                await conn.execute("""
                    UPDATE prod_registro_requerimiento_mp
                    SET cantidad_consumida = GREATEST(cantidad_consumida - $1, 0), updated_at = CURRENT_TIMESTAMP
                    WHERE registro_id = $2 AND item_id = $3 AND talla_id IS NULL
                """, float(salida['cantidad']), salida['registro_id'], salida['item_id'])
        return {"message": "Salida eliminada y stock restaurado"}

# ==================== ENDPOINTS ROLLOS ====================

@router.get("/inventario-rollos")
async def get_rollos(item_id: str = None, activo: bool = None):
    pool = await get_pool()
    async with pool.acquire() as conn:
        query = "SELECT * FROM prod_inventario_rollos WHERE 1=1"
        params = []
        if item_id:
            params.append(item_id)
            query += f" AND item_id = ${len(params)}"
        if activo is not None:
            params.append(activo)
            query += f" AND activo = ${len(params)}"
            if activo:
                query += " AND metraje_disponible > 0"
        query += " ORDER BY created_at DESC"
        rows = await conn.fetch(query, *params)
        result = []
        for r in rows:
            d = row_to_dict(r)
            item = await conn.fetchrow("SELECT nombre, codigo FROM prod_inventario WHERE id = $1", d.get('item_id'))
            d['item_nombre'] = item['nombre'] if item else ""
            d['item_codigo'] = item['codigo'] if item else ""
            result.append(d)
        return result

@router.get("/inventario-rollos/{rollo_id}")
async def get_rollo(rollo_id: str):
    pool = await get_pool()
    async with pool.acquire() as conn:
        rollo = await conn.fetchrow("SELECT * FROM prod_inventario_rollos WHERE id = $1", rollo_id)
        if not rollo:
            raise HTTPException(status_code=404, detail="Rollo no encontrado")
        d = row_to_dict(rollo)
        item = await conn.fetchrow("SELECT nombre, codigo FROM prod_inventario WHERE id = $1", d.get('item_id'))
        d['item_nombre'] = item['nombre'] if item else ""
        d['item_codigo'] = item['codigo'] if item else ""
        return d

# ==================== ENDPOINTS AJUSTES INVENTARIO ====================

@router.get("/inventario-ajustes")
async def get_ajustes():
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch("SELECT * FROM prod_inventario_ajustes ORDER BY fecha DESC")
        result = []
        for r in rows:
            d = row_to_dict(r)
            item = await conn.fetchrow("SELECT nombre, codigo, control_por_rollos FROM prod_inventario WHERE id = $1", d.get('item_id'))
            d['item_nombre'] = item['nombre'] if item else ""
            d['item_codigo'] = item['codigo'] if item else ""
            d['control_por_rollos'] = item['control_por_rollos'] if item else False
            # Info del rollo si existe
            if d.get('rollo_id'):
                rollo = await conn.fetchrow("SELECT numero_rollo, tono FROM prod_inventario_rollos WHERE id = $1", d['rollo_id'])
                d['numero_rollo'] = rollo['numero_rollo'] if rollo else ""
                d['tono'] = rollo['tono'] if rollo else ""
            result.append(d)
        return result

@router.post("/inventario-ajustes")
async def create_ajuste(input: AjusteInventarioCreate, current_user: dict = Depends(get_current_user)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        item = await conn.fetchrow("SELECT * FROM prod_inventario WHERE id = $1", input.item_id)
        if not item:
            raise HTTPException(status_code=404, detail="Item de inventario no encontrado")
        if input.tipo not in ["entrada", "salida"]:
            raise HTTPException(status_code=400, detail="Tipo debe ser 'entrada' o 'salida'")
        
        control_por_rollos = item['control_por_rollos']
        
        # Validaciones para items con control por rollos
        if control_por_rollos:
            if input.tipo == "salida":
                # Para salida de rollo, necesitamos el rollo_id OBLIGATORIO
                if not input.rollo_id:
                    raise HTTPException(status_code=400, detail="Este item requiere seleccionar un rollo para ajuste de salida")
                rollo = await conn.fetchrow("SELECT * FROM prod_inventario_rollos WHERE id = $1", input.rollo_id)
                if not rollo:
                    raise HTTPException(status_code=404, detail="Rollo no encontrado")
                if rollo['item_id'] != input.item_id:
                    raise HTTPException(status_code=400, detail="El rollo no pertenece a este item")
                if float(rollo['metraje_disponible']) < input.cantidad:
                    raise HTTPException(status_code=400, detail=f"Metraje insuficiente en rollo. Disponible: {rollo['metraje_disponible']}")
            elif input.tipo == "entrada" and input.rollo_id:
                # Entrada con rollo específico (opcional) - aumentar metraje del rollo
                rollo = await conn.fetchrow("SELECT * FROM prod_inventario_rollos WHERE id = $1", input.rollo_id)
                if not rollo:
                    raise HTTPException(status_code=404, detail="Rollo no encontrado")
                if rollo['item_id'] != input.item_id:
                    raise HTTPException(status_code=400, detail="El rollo no pertenece a este item")
            # Si es entrada sin rollo_id, solo se aumenta el stock general (permitido)
        else:
            # Items sin control por rollos no deben tener rollo_id
            if input.rollo_id:
                raise HTTPException(status_code=400, detail="Este item no usa control por rollos")
            if input.tipo == "salida" and float(item['stock_actual']) < input.cantidad:
                raise HTTPException(status_code=400, detail=f"Stock insuficiente. Disponible: {item['stock_actual']}")
        
        ajuste = AjusteInventario(**input.model_dump())
        await conn.execute(
            """INSERT INTO prod_inventario_ajustes (id, item_id, tipo, cantidad, motivo, observaciones, rollo_id, fecha)
               VALUES ($1,$2,$3,$4,$5,$6,$7,$8)""",
            ajuste.id, ajuste.item_id, ajuste.tipo, ajuste.cantidad, ajuste.motivo, ajuste.observaciones, ajuste.rollo_id, ajuste.fecha.replace(tzinfo=None)
        )
        
        # Actualizar stock del item
        incremento = input.cantidad if input.tipo == "entrada" else -input.cantidad
        await conn.execute("UPDATE prod_inventario SET stock_actual = stock_actual + $1 WHERE id = $2", incremento, input.item_id)
        
        # Actualizar metraje del rollo si aplica (solo si hay rollo_id)
        if control_por_rollos and input.rollo_id:
            rollo = await conn.fetchrow("SELECT ingreso_id FROM prod_inventario_rollos WHERE id = $1", input.rollo_id)
            if rollo:
                if input.tipo == "entrada":
                    await conn.execute("UPDATE prod_inventario_rollos SET metraje_disponible = metraje_disponible + $1, metraje = metraje + $1 WHERE id = $2", input.cantidad, input.rollo_id)
                    await conn.execute("UPDATE prod_inventario_ingresos SET cantidad_disponible = cantidad_disponible + $1, cantidad = cantidad + $1 WHERE id = $2", input.cantidad, rollo['ingreso_id'])
                else:  # salida
                    await conn.execute("UPDATE prod_inventario_rollos SET metraje_disponible = metraje_disponible - $1 WHERE id = $2", input.cantidad, input.rollo_id)
                    await conn.execute("UPDATE prod_inventario_ingresos SET cantidad_disponible = cantidad_disponible - $1 WHERE id = $2", input.cantidad, rollo['ingreso_id'])
        
        stock_antes = float(item['stock_actual'])
        stock_despues = stock_antes + incremento
        await audit_log_safe(conn, get_usuario(current_user), "UPDATE", "inventario", "prod_inventario_ajustes", ajuste.id,
            datos_antes={"stock_actual": stock_antes},
            datos_despues={"stock_actual": stock_despues, "tipo": input.tipo, "cantidad": input.cantidad, "motivo": input.motivo})
    await registrar_actividad(pool, current_user['id'], current_user.get('username', ''), "editar",
        tabla_afectada="inventario", registro_id=ajuste.id, registro_nombre=item['nombre'],
        descripcion=f"Ajuste {input.tipo} de {input.cantidad} uds en {item['nombre']}: {input.motivo}")
    return ajuste

class AjusteUpdateData(BaseModel):
    motivo: str = ""
    observaciones: str = ""

@router.put("/inventario-ajustes/{ajuste_id}")
async def update_ajuste(ajuste_id: str, input: AjusteUpdateData, _u=Depends(get_current_user)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        ajuste = await conn.fetchrow("SELECT * FROM prod_inventario_ajustes WHERE id = $1", ajuste_id)
        if not ajuste:
            raise HTTPException(status_code=404, detail="Ajuste no encontrado")
        await conn.execute("UPDATE prod_inventario_ajustes SET motivo=$1, observaciones=$2 WHERE id=$3", 
                          input.motivo, input.observaciones, ajuste_id)
        return {"message": "Ajuste actualizado"}

@router.delete("/inventario-ajustes/{ajuste_id}")
async def delete_ajuste(ajuste_id: str, _u=Depends(get_current_user)):
    pool = await get_pool()
    async with pool.acquire() as conn:
        ajuste = await conn.fetchrow("SELECT * FROM prod_inventario_ajustes WHERE id = $1", ajuste_id)
        if not ajuste:
            raise HTTPException(status_code=404, detail="Ajuste no encontrado")

        item = await conn.fetchrow("SELECT control_por_rollos FROM prod_inventario WHERE id = $1", ajuste['item_id'])

        incremento = -float(ajuste['cantidad']) if ajuste['tipo'] == "entrada" else float(ajuste['cantidad'])
        # Validar stock negativo solo si NO es cascade de ajuste_migracion
        # (ese cascade elimina las salidas y restaura stock, compensando el decremento)
        if ajuste['tipo'] == "entrada" and ajuste.get('subtipo') != 'ajuste_migracion':
            current_item = await conn.fetchrow("SELECT stock_actual FROM prod_inventario WHERE id = $1", ajuste['item_id'])
            if current_item and float(current_item['stock_actual']) < float(ajuste['cantidad']):
                raise HTTPException(status_code=400, detail="No se puede eliminar: dejaría el stock negativo")

        async with conn.transaction():
            # Caso especial: ajuste de reversión de modo carga inicial
            # → eliminar en CASCADA las salidas que este ajuste revertía.
            # Al desactivar carga inicial pasaron 2 cosas:
            #   (a) ajuste +Q sumó stock  (b) se restauró +Q en cantidad_disponible FIFO
            # Al eliminar el ajuste con cascada debemos deshacer (a)+(b) y además eliminar la salida.
            # Como la salida original ya había restado stock y FIFO, el resultado neto
            # es volver al estado previo a la carga inicial (como si la salida nunca hubiera existido).
            if ajuste.get('subtipo') == 'ajuste_migracion':
                salidas_revertidas = await conn.fetch(
                    "SELECT id, cantidad, detalle_fifo, registro_id, talla_id, item_id FROM prod_inventario_salidas WHERE revertida_por_migracion_id = $1",
                    ajuste_id,
                )
                # Deshacer la restauración de FIFO que hizo el desactivar (resta las cantidades a cada capa)
                for sal in salidas_revertidas:
                    detalle = sal['detalle_fifo']
                    if isinstance(detalle, str):
                        try:
                            detalle = json.loads(detalle)
                        except Exception:
                            detalle = []
                    for capa in (detalle or []):
                        ingreso_id = capa.get('ingreso_id')
                        cant = capa.get('cantidad')
                        if ingreso_id and cant:
                            await conn.execute(
                                "UPDATE prod_inventario_ingresos SET cantidad_disponible = cantidad_disponible - $1 WHERE id = $2",
                                float(cant), ingreso_id,
                            )
                # Luego eliminar cada salida y restaurar stock + FIFO como lo hace el DELETE normal
                await conn.execute(
                    "UPDATE prod_inventario_salidas SET revertida_por_migracion_id = NULL WHERE revertida_por_migracion_id = $1",
                    ajuste_id,
                )
                for sal in salidas_revertidas:
                    detalle = sal['detalle_fifo']
                    if isinstance(detalle, str):
                        try:
                            detalle = json.loads(detalle)
                        except Exception:
                            detalle = []
                    # Restaurar capas FIFO que consumió la salida original
                    for capa in (detalle or []):
                        ingreso_id = capa.get('ingreso_id')
                        cant = capa.get('cantidad')
                        if ingreso_id and cant:
                            await conn.execute(
                                "UPDATE prod_inventario_ingresos SET cantidad_disponible = cantidad_disponible + $1 WHERE id = $2",
                                float(cant), ingreso_id,
                            )
                    # Restaurar stock_actual
                    await conn.execute(
                        "UPDATE prod_inventario SET stock_actual = stock_actual + $1 WHERE id = $2",
                        float(sal['cantidad']), ajuste['item_id'],
                    )
                    # Decrementar cantidad_consumida en el requerimiento del registro
                    if sal['registro_id']:
                        if sal.get('talla_id'):
                            await conn.execute("""
                                UPDATE prod_registro_requerimiento_mp
                                SET cantidad_consumida = GREATEST(cantidad_consumida - $1, 0),
                                    updated_at = CURRENT_TIMESTAMP
                                WHERE registro_id = $2 AND item_id = $3 AND talla_id = $4
                            """, float(sal['cantidad']), sal['registro_id'], sal['item_id'], sal['talla_id'])
                        else:
                            await conn.execute("""
                                UPDATE prod_registro_requerimiento_mp
                                SET cantidad_consumida = GREATEST(cantidad_consumida - $1, 0),
                                    updated_at = CURRENT_TIMESTAMP
                                WHERE registro_id = $2 AND item_id = $3 AND talla_id IS NULL
                            """, float(sal['cantidad']), sal['registro_id'], sal['item_id'])
                    # Eliminar la salida
                    await conn.execute("DELETE FROM prod_inventario_salidas WHERE id = $1", sal['id'])

            await conn.execute("DELETE FROM prod_inventario_ajustes WHERE id = $1", ajuste_id)
            await conn.execute("UPDATE prod_inventario SET stock_actual = stock_actual + $1 WHERE id = $2", incremento, ajuste['item_id'])

            # Revertir metraje del rollo si aplica
            if item and item['control_por_rollos'] and ajuste.get('rollo_id'):
                rollo = await conn.fetchrow("SELECT ingreso_id FROM prod_inventario_rollos WHERE id = $1", ajuste['rollo_id'])
                if rollo:
                    if ajuste['tipo'] == "entrada":
                        # Revertir entrada = restar metraje
                        await conn.execute("UPDATE prod_inventario_rollos SET metraje_disponible = metraje_disponible - $1, metraje = metraje - $1 WHERE id = $2", float(ajuste['cantidad']), ajuste['rollo_id'])


# ==================== DISPONIBILIDAD Y TIPOS ====================

@router.get("/inventario/{item_id}/disponibilidad")
async def get_disponibilidad(item_id: str, current_user: dict = Depends(get_current_user)):
    """Obtiene disponibilidad real (stock - reservas)"""
    pool = await get_pool()
    async with pool.acquire() as conn:
        item = await conn.fetchrow("SELECT * FROM prod_inventario WHERE id = $1", item_id)
        if not item:
            raise HTTPException(status_code=404, detail="Item no encontrado")
        stock_actual = float(item['stock_actual'] or 0)
        total_reservado = await conn.fetchval("""
            SELECT COALESCE(SUM(rl.cantidad_reservada - rl.cantidad_liberada), 0)
            FROM prod_inventario_reservas_linea rl
            JOIN prod_inventario_reservas r ON rl.reserva_id = r.id
            WHERE rl.item_id = $1 AND r.estado = 'ACTIVA'
        """, item_id)
        total_reservado = float(total_reservado or 0)
        return {
            "item_id": item_id,
            "codigo": item['codigo'],
            "nombre": item['nombre'],
            "tipo_item": item.get('tipo_item', 'MP'),
            "stock_actual": stock_actual,
            "total_reservado": total_reservado,
            "disponible": max(0, stock_actual - total_reservado),
            "control_por_rollos": item['control_por_rollos']
        }


@router.get("/inventario-tipos")
async def get_tipos_item():
    """Lista tipos de item válidos"""
    return {
        "tipos": [
            {"codigo": "MP", "nombre": "Materia Prima", "descripcion": "Telas y materiales principales"},
            {"codigo": "AVIO", "nombre": "Avío", "descripcion": "Botones, cierres, etiquetas, etc."},
            {"codigo": "SERVICIO", "nombre": "Servicio", "descripcion": "Servicios externos (no genera stock)"},
            {"codigo": "PT", "nombre": "Producto Terminado", "descripcion": "Prendas terminadas"}
        ]
    }
