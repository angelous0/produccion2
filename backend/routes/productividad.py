"""
Reporte de Productividad por rango de fechas.

Permite controlar cuánto está produciendo cada servicio (Corte, Costura, Acabado,
Lavandería, etc.) y cada persona (taller/operario) en un período específico.

Endpoints:
  GET /api/reportes/productividad
       ?fecha_desde=YYYY-MM-DD&fecha_hasta=YYYY-MM-DD
       &servicio_id=...&persona_id=...&estado=...
       Devuelve: resumen general, por servicio, por persona, y lista de movimientos.

  GET /api/reportes/productividad/export
       Misma firma, devuelve Excel con 4 hojas (Resumen, Por Servicio, Por Persona, Detalle).

Métricas:
  - cantidad_movimientos: # de movimientos completados (fecha_fin presente)
  - cantidad_enviada: total de prendas que entraron al servicio
  - cantidad_recibida: total que salieron OK
  - diferencia: cantidad_enviada - cantidad_recibida (mermas / fallados)
  - costo_total: suma de costo_calculado
  - registros_distintos: # de lotes diferentes trabajados
"""
from io import BytesIO
from datetime import date, datetime
from typing import Optional
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse

from db import get_pool
from auth_utils import get_current_user

router = APIRouter(prefix="/api")


def _build_filtros(
    fecha_desde: Optional[date],
    fecha_hasta: Optional[date],
    servicio_id: Optional[str],
    persona_id: Optional[str],
):
    conds = ["m.fecha_fin IS NOT NULL"]  # Solo movimientos completados
    params: list = []
    if fecha_desde:
        params.append(fecha_desde); conds.append(f"m.fecha_fin >= ${len(params)}")
    if fecha_hasta:
        params.append(fecha_hasta); conds.append(f"m.fecha_fin <= ${len(params)}")
    if servicio_id:
        params.append(servicio_id); conds.append(f"m.servicio_id = ${len(params)}")
    if persona_id:
        params.append(persona_id); conds.append(f"m.persona_id = ${len(params)}")
    return " AND ".join(conds), params


@router.get("/reportes/productividad")
async def reporte_productividad(
    fecha_desde: Optional[date] = None,
    fecha_hasta: Optional[date] = None,
    servicio_id: Optional[str] = None,
    persona_id: Optional[str] = None,
    incluir_detalle: bool = Query(True, description="Si false, omite la lista de detalle (más rápido)"),
    current_user: dict = Depends(get_current_user),
):
    where, params = _build_filtros(fecha_desde, fecha_hasta, servicio_id, persona_id)
    # Filtros para cascada (sin servicio/persona)
    where_cascada_serv, params_cascada_serv = _build_filtros(fecha_desde, fecha_hasta, None, persona_id)
    where_cascada_pers, params_cascada_pers = _build_filtros(fecha_desde, fecha_hasta, servicio_id, None)

    pool = await get_pool()
    async with pool.acquire() as conn:
        # Una sola CTE para evitar 3 consultas con el mismo WHERE
        base_query = f"""
            WITH movs AS (
                SELECT m.* FROM produccion.prod_movimientos_produccion m WHERE {where}
            )
            SELECT
              -- Resumen
              (SELECT COUNT(*) FROM movs)                                      AS r_movimientos,
              (SELECT COALESCE(SUM(cantidad_enviada),0) FROM movs)             AS r_enviadas,
              (SELECT COALESCE(SUM(cantidad_recibida),0) FROM movs)            AS r_recibidas,
              (SELECT COALESCE(SUM(cantidad_enviada - COALESCE(cantidad_recibida,0)),0) FROM movs) AS r_diferencia,
              (SELECT COALESCE(SUM(costo_calculado),0) FROM movs)              AS r_costo,
              (SELECT COUNT(DISTINCT registro_id) FROM movs)                   AS r_registros,
              (SELECT COUNT(DISTINCT persona_id) FROM movs)                    AS r_personas,
              (SELECT COUNT(DISTINCT servicio_id) FROM movs)                   AS r_servicios
        """
        resumen = await conn.fetchrow(base_query, *params)

        # Por servicio (con tabla servicios para join)
        por_servicio = await conn.fetch(f"""
            SELECT
                s.id AS servicio_id, s.nombre AS servicio_nombre,
                COUNT(*) AS movimientos,
                COALESCE(SUM(m.cantidad_enviada),0)::int AS enviadas,
                COALESCE(SUM(m.cantidad_recibida),0)::int AS recibidas,
                COALESCE(SUM(m.cantidad_enviada - COALESCE(m.cantidad_recibida,0)),0)::int AS diferencia,
                COALESCE(SUM(m.costo_calculado),0)::numeric AS costo_total,
                COUNT(DISTINCT m.registro_id)::int AS registros_distintos,
                COUNT(DISTINCT m.persona_id)::int AS personas_distintas
              FROM produccion.prod_movimientos_produccion m
              LEFT JOIN produccion.prod_servicios_produccion s ON m.servicio_id = s.id
             WHERE {where}
             GROUP BY s.id, s.nombre
             ORDER BY costo_total DESC, movimientos DESC
        """, *params)

        # Por persona
        por_persona = await conn.fetch(f"""
            SELECT
                p.id AS persona_id, p.nombre AS persona_nombre, p.tipo AS persona_tipo,
                COUNT(*) AS movimientos,
                COALESCE(SUM(m.cantidad_enviada),0)::int AS enviadas,
                COALESCE(SUM(m.cantidad_recibida),0)::int AS recibidas,
                COALESCE(SUM(m.cantidad_enviada - COALESCE(m.cantidad_recibida,0)),0)::int AS diferencia,
                COALESCE(SUM(m.costo_calculado),0)::numeric AS costo_total,
                COUNT(DISTINCT m.registro_id)::int AS registros_distintos,
                COUNT(DISTINCT m.servicio_id)::int AS servicios_distintos
              FROM produccion.prod_movimientos_produccion m
              LEFT JOIN produccion.prod_personas_produccion p ON m.persona_id = p.id
             WHERE {where}
             GROUP BY p.id, p.nombre, p.tipo
             ORDER BY costo_total DESC, movimientos DESC
        """, *params)

        # Cascada: servicios y personas DISPONIBLES en el período (con el otro filtro aplicado)
        servicios_disponibles = await conn.fetch(f"""
            SELECT DISTINCT s.id, s.nombre
              FROM produccion.prod_movimientos_produccion m
              JOIN produccion.prod_servicios_produccion s ON m.servicio_id = s.id
             WHERE {where_cascada_serv}
             ORDER BY s.nombre
        """, *params_cascada_serv)

        personas_disponibles = await conn.fetch(f"""
            SELECT DISTINCT p.id, p.nombre, p.tipo
              FROM produccion.prod_movimientos_produccion m
              JOIN produccion.prod_personas_produccion p ON m.persona_id = p.id
             WHERE {where_cascada_pers}
             ORDER BY p.nombre
        """, *params_cascada_pers)

        # Detalle: solo si lo piden (el dashboard inicial puede omitirlo para velocidad)
        detalle = []
        if incluir_detalle:
            # IMPORTANTE: la mayoría de registros usan modelo_manual (JSONB) en lugar
            # de modelo_id (FK). Para resolver tipo/marca/entalle hacemos JOIN usando
            # COALESCE entre el FK del modelo y el id que viene en el JSONB.
            detalle = await conn.fetch(f"""
                SELECT
                    m.id, m.fecha_inicio, m.fecha_fin,
                    m.cantidad_enviada, m.cantidad_recibida,
                    (m.cantidad_enviada - COALESCE(m.cantidad_recibida, 0)) AS diferencia,
                    m.costo_calculado, m.tarifa_aplicada,
                    r.n_corte, r.estado AS registro_estado,
                    COALESCE(mo.nombre, r.modelo_manual->>'nombre_modelo') AS modelo_nombre,
                    t.nombre AS tipo_nombre,
                    ma.nombre AS marca_nombre,
                    en.nombre AS entalle_nombre,
                    s.nombre AS servicio_nombre,
                    p.nombre AS persona_nombre,
                    p.tipo AS persona_tipo
                  FROM produccion.prod_movimientos_produccion m
                  LEFT JOIN produccion.prod_registros r ON m.registro_id = r.id
                  LEFT JOIN produccion.prod_modelos mo ON r.modelo_id = mo.id
                  LEFT JOIN produccion.prod_tipos t
                    ON t.id = COALESCE(mo.tipo_id, r.modelo_manual->>'tipo_id')
                  LEFT JOIN produccion.prod_marcas ma
                    ON ma.id = COALESCE(mo.marca_id, r.modelo_manual->>'marca_id')
                  LEFT JOIN produccion.prod_entalles en
                    ON en.id = COALESCE(mo.entalle_id, r.modelo_manual->>'entalle_id')
                  LEFT JOIN produccion.prod_servicios_produccion s ON m.servicio_id = s.id
                  LEFT JOIN produccion.prod_personas_produccion p ON m.persona_id = p.id
                 WHERE {where}
                 ORDER BY m.fecha_fin DESC
                 LIMIT 500
            """, *params)

    def _f(v): return float(v or 0)
    def _i(v): return int(v or 0)

    return {
        "filtros": {
            "fecha_desde": fecha_desde.isoformat() if fecha_desde else None,
            "fecha_hasta": fecha_hasta.isoformat() if fecha_hasta else None,
            "servicio_id": servicio_id,
            "persona_id": persona_id,
        },
        "resumen": {
            "movimientos": _i(resumen["r_movimientos"]),
            "enviadas": _i(resumen["r_enviadas"]),
            "recibidas": _i(resumen["r_recibidas"]),
            "diferencia": _i(resumen["r_diferencia"]),
            "costo_total": round(_f(resumen["r_costo"]), 2),
            "registros_distintos": _i(resumen["r_registros"]),
            "personas_distintas": _i(resumen["r_personas"]),
            "servicios_distintos": _i(resumen["r_servicios"]),
        },
        "cascada": {
            "servicios_disponibles": [
                {"id": r["id"], "nombre": r["nombre"]} for r in servicios_disponibles
            ],
            "personas_disponibles": [
                {"id": r["id"], "nombre": r["nombre"], "tipo": r["tipo"]}
                for r in personas_disponibles
            ],
        },
        "por_servicio": [{
            "servicio_id": r["servicio_id"],
            "servicio_nombre": r["servicio_nombre"] or "(sin servicio)",
            "movimientos": _i(r["movimientos"]),
            "enviadas": _i(r["enviadas"]),
            "recibidas": _i(r["recibidas"]),
            "diferencia": _i(r["diferencia"]),
            "costo_total": round(_f(r["costo_total"]), 2),
            "registros_distintos": _i(r["registros_distintos"]),
            "personas_distintas": _i(r["personas_distintas"]),
        } for r in por_servicio],
        "por_persona": [{
            "persona_id": r["persona_id"],
            "persona_nombre": r["persona_nombre"] or "(sin persona)",
            "persona_tipo": r["persona_tipo"],
            "movimientos": _i(r["movimientos"]),
            "enviadas": _i(r["enviadas"]),
            "recibidas": _i(r["recibidas"]),
            "diferencia": _i(r["diferencia"]),
            "costo_total": round(_f(r["costo_total"]), 2),
            "registros_distintos": _i(r["registros_distintos"]),
            "servicios_distintos": _i(r["servicios_distintos"]),
        } for r in por_persona],
        "detalle": [{
            "id": r["id"],
            "fecha_inicio": r["fecha_inicio"].isoformat() if r["fecha_inicio"] else None,
            "fecha_fin": r["fecha_fin"].isoformat() if r["fecha_fin"] else None,
            "n_corte": r["n_corte"],
            "modelo_nombre": r["modelo_nombre"],
            "tipo_nombre": r["tipo_nombre"],
            "marca_nombre": r["marca_nombre"],
            "entalle_nombre": r["entalle_nombre"],
            "registro_estado": r["registro_estado"],
            "servicio_nombre": r["servicio_nombre"] or "(sin servicio)",
            "persona_nombre": r["persona_nombre"] or "(sin persona)",
            "persona_tipo": r["persona_tipo"],
            "cantidad_enviada": _i(r["cantidad_enviada"]),
            "cantidad_recibida": _i(r["cantidad_recibida"]),
            "diferencia": _i(r["diferencia"]),
            "costo_calculado": round(_f(r["costo_calculado"]), 2),
            "tarifa_aplicada": round(_f(r["tarifa_aplicada"]), 2),
        } for r in detalle],
    }


@router.get("/reportes/productividad/export")
async def exportar_productividad(
    fecha_desde: Optional[date] = None,
    fecha_hasta: Optional[date] = None,
    servicio_id: Optional[str] = None,
    persona_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    """Exporta reporte de productividad a Excel (4 hojas)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    data = await reporte_productividad(
        fecha_desde, fecha_hasta, servicio_id, persona_id, current_user)

    wb = Workbook()

    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    def write_sheet(ws, headers, rows):
        for col_idx, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=col_idx, value=h)
            c.fill = header_fill; c.font = header_font
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[c.column_letter].width = max(15, len(h) + 2)
        for row_idx, r in enumerate(rows, start=2):
            for col_idx, v in enumerate(r, start=1):
                ws.cell(row=row_idx, column=col_idx, value=v)

    # Hoja 1: Resumen
    ws = wb.active
    ws.title = "Resumen"
    res = data["resumen"]
    f = data["filtros"]
    ws.append(["REPORTE DE PRODUCTIVIDAD"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Período:", f"{f['fecha_desde'] or '-'}  a  {f['fecha_hasta'] or '-'}"])
    ws.append([])
    ws.append(["Métrica", "Valor"])
    ws.cell(row=5, column=1).font = header_font; ws.cell(row=5, column=1).fill = header_fill
    ws.cell(row=5, column=2).font = header_font; ws.cell(row=5, column=2).fill = header_fill
    items = [
        ("Movimientos completados", res["movimientos"]),
        ("Prendas enviadas", res["enviadas"]),
        ("Prendas recibidas (OK)", res["recibidas"]),
        ("Diferencia (mermas / fallados)", res["diferencia"]),
        ("Costo total (S/.)", res["costo_total"]),
        ("Lotes distintos trabajados", res["registros_distintos"]),
        ("Personas/talleres activos", res["personas_distintas"]),
        ("Servicios distintos", res["servicios_distintos"]),
    ]
    for k, v in items:
        ws.append([k, v])
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 18

    # Hoja 2: Por Servicio
    ws2 = wb.create_sheet("Por Servicio")
    write_sheet(ws2, [
        "Servicio", "Movimientos", "Enviadas", "Recibidas", "Diferencia",
        "Costo total (S/.)", "Lotes", "Personas",
    ], [[
        r["servicio_nombre"], r["movimientos"], r["enviadas"], r["recibidas"],
        r["diferencia"], r["costo_total"], r["registros_distintos"], r["personas_distintas"],
    ] for r in data["por_servicio"]])

    # Hoja 3: Por Persona
    ws3 = wb.create_sheet("Por Persona")
    write_sheet(ws3, [
        "Persona/Taller", "Tipo", "Movimientos", "Enviadas", "Recibidas",
        "Diferencia", "Costo total (S/.)", "Lotes", "Servicios",
    ], [[
        r["persona_nombre"], r["persona_tipo"] or "", r["movimientos"], r["enviadas"],
        r["recibidas"], r["diferencia"], r["costo_total"],
        r["registros_distintos"], r["servicios_distintos"],
    ] for r in data["por_persona"]])

    # Hoja 4: Detalle
    ws4 = wb.create_sheet("Detalle")
    write_sheet(ws4, [
        "Fecha fin", "N° Corte", "Modelo", "Tipo", "Marca", "Entalle", "Estado lote",
        "Servicio", "Persona/Taller", "Tipo persona",
        "Enviadas", "Recibidas", "Dif.", "Tarifa", "Costo",
    ], [[
        r["fecha_fin"], r["n_corte"], r["modelo_nombre"] or "",
        r.get("tipo_nombre") or "", r.get("marca_nombre") or "", r.get("entalle_nombre") or "",
        r["registro_estado"] or "",
        r["servicio_nombre"], r["persona_nombre"], r["persona_tipo"] or "",
        r["cantidad_enviada"], r["cantidad_recibida"], r["diferencia"],
        r["tarifa_aplicada"], r["costo_calculado"],
    ] for r in data["detalle"]])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    nombre = "productividad"
    if fecha_desde and fecha_hasta:
        nombre += f"_{fecha_desde}_a_{fecha_hasta}"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre}.xlsx"'},
    )
